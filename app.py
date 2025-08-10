import concurrent.futures
import uuid
import os
import json
import tempfile
import re
import time
import mimetypes

import google.generativeai as genai
import gspread
import streamlit as st

from google.oauth2.service_account import Credentials
from openpyxl.utils import get_column_letter

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

DEFAULT_SHEET_ID = "17tMHStXQYXaIQHQIA4jdUyHaYt_tuoNCEEuJCstWEuw"

COMPANY_NAME_ROW = 1
CONTACT_INFO_ROW = 2
HEADER_ROW = 3
ITEM_MASTER_LIST_COL = 2
COLUMNS_PER_SUPPLIER = 4

SUMMARY_LABELS = [
    "รวมเป็นเงิน",
    "ภาษีมูลค่าเพิ่ม 7%",
    "ยอดรวมทั้งสิ้น",
    "กำหนดยืนราคา (วัน)",
    "ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO",
    "การชำระเงิน",
    "อื่น ๆ",
]

def extract_sheet_id_from_url(url):
    if not url:
        return None
    if "/" not in url and " " not in url and len(url) > 20:
        return url
    m = re.search(r"spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None

def extract_json_from_text(text):
    if not text:
        return None
    blocks = re.findall(r"```(?:json)?\s*(\[[\s\S]*?\]|\{[\s\S]*?\})\s*```", text)
    if blocks:
        candidates = blocks
    else:
        start_brace = text.find("{")
        start_bracket = text.find("[")
        if start_bracket != -1 and (start_bracket < start_brace or start_brace == -1):
            start = start_bracket
            end = text.rfind("]") + 1
        elif start_brace != -1:
            start = start_brace
            end = text.rfind("}") + 1
        else:
            return None
        candidates = [text[start:end]] if end > start else []
    for cand in candidates:
        json_str = cand
        cleaned_json = re.sub(r",\s*([}\]])", r"\1", json_str)
        try:
            return json.loads(cleaned_json)
        except json.JSONDecodeError:
            continue
    return None

def extract_contact_info(text):
    if not text:
        return ""
    phone_pattern = r"(?<!\w)((0\d{1,2}[-\s]?\d{3}[-\s]?\d{3,4})|(0\d{2}[-\s]?\d{7})|(0\d{2}[-\s]?\d{3}[-\s]?\d{4}))(?!\w)"
    email_pattern = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
    phone_matches = re.findall(phone_pattern, text)
    email_matches = re.findall(email_pattern, text)
    phone_numbers = [m[0] for m in phone_matches] if phone_matches else []
    emails = sorted(set(email_matches))
    phones = []
    for phone in phone_numbers:
        clean_phone = re.sub(r"\s", "", phone)
        if len(clean_phone) >= 9:
            phones.append(clean_phone)
    phones = sorted(set(phones))
    contact_parts = []
    if emails:
        contact_parts.append(f"Email: {', '.join(emails)}")
    if phones:
        contact_parts.append(f"Phone: {', '.join(phones)}")
    return ", ".join(contact_parts)

def sanitize_filename_for_upload(file_name: str) -> str:
    if not file_name:
        return f"upload_{uuid.uuid4().hex}.file"
    base_name, extension = os.path.splitext(file_name)
    safe_base_name = re.sub(r'[^a-zA-Z0-9._-]', '_', base_name)
    safe_base_name = re.sub(r'__+', '_', safe_base_name)
    safe_base_name = safe_base_name.strip('_')
    if not safe_base_name:
        safe_base_name = f"file_{uuid.uuid4().hex}"
    return f"{safe_base_name}{extension}"

def clean_product_name(name):
    if not name:
        return "Unknown Product"
    return re.sub(r"^\s*\d+[\.\)\-]\s*", "", name.strip())

def _to_number_or_default(val, default):
    s = str(val)
    s2 = s.replace(",", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", s2):
        try:
            return float(s2)
        except Exception:
            return default
    return default

def validate_json_data(json_data):
    if not json_data:
        return {
            "company": "Unknown Company",
            "contact": "",
            "vat": False,
            "products": [],
            "totalPrice": 0,
            "totalVat": 0,
            "totalPriceIncludeVat": 0,
            "priceGuaranteeDay": 0,
            "deliveryTime": "",
            "paymentTerms": "",
            "otherNotes": "",
        }
    if not json_data.get("company"):
        json_data["company"] = "Unknown Company"

    if "contact" in json_data:
        if isinstance(json_data["contact"], dict):
            contact_parts = []
            if "email" in json_data["contact"] and json_data["contact"]["email"]:
                contact_parts.append(f"Email: {json_data['contact']['email']}")
            if "phone" in json_data["contact"] and json_data["contact"]["phone"]:
                contact_parts.append(f"Phone: {json_data['contact']['phone']}")
            json_data["contact"] = ", ".join(contact_parts)
        else:
            json_data["contact"] = extract_contact_info(str(json_data["contact"]))
    else:
        json_data["contact"] = ""

    if "vat" not in json_data:
        json_data["vat"] = False
    else:
        json_data["vat"] = bool(json_data["vat"])

    if not json_data.get("products"):
        json_data["products"] = []
    for product in json_data.get("products", []):
        product["name"] = clean_product_name(product.get("name") or "Unknown Product")
        product["quantity"] = _to_number_or_default(product.get("quantity", 1), 1)
        if product["quantity"] <= 0:
            product["quantity"] = 1
        product["unit"] = product.get("unit") or "ชิ้น"
        product["pricePerUnit"] = _to_number_or_default(product.get("pricePerUnit", 0), 0)
        provided_total = _to_number_or_default(product.get("totalPrice", None), None)
        if provided_total is None:
            product["totalPrice"] = round(product["quantity"] * product["pricePerUnit"], 2)
        else:
            product["totalPrice"] = provided_total

    computed_total = sum(p.get("totalPrice", 0) for p in json_data.get("products", []))
    json_data["totalPrice"] = _to_number_or_default(json_data.get("totalPrice", computed_total), computed_total)
    json_data["totalVat"] = _to_number_or_default(json_data.get("totalVat", round(json_data["totalPrice"] * 0.07, 2)), round(json_data["totalPrice"] * 0.07, 2))
    json_data["totalPriceIncludeVat"] = _to_number_or_default(
        json_data.get("totalPriceIncludeVat", round(json_data["totalPrice"] + json_data["totalVat"], 2)),
        round(json_data["totalPrice"] + json_data["totalVat"], 2),
    )

    if "priceGuaranteeDay" not in json_data:
        json_data["priceGuaranteeDay"] = 0
    if "deliveryTime" not in json_data:
        json_data["deliveryTime"] = ""
    if "paymentTerms" not in json_data:
        json_data["paymentTerms"] = ""
    if "otherNotes" not in json_data:
        json_data["otherNotes"] = ""

    return json_data

def enhance_with_gemini(json_data: dict, prompt_template: str):
    if not json_data:
        return None
    validation_formatted = prompt_template.format(
        extracted_json=json.dumps(json_data, ensure_ascii=False)
    )
    model = genai.GenerativeModel(model_name="gemini-2.5-flash")
    response = model.generate_content(validation_formatted)
    enhanced_text = (response.text or "").strip()
    enhanced = extract_json_from_text(enhanced_text) or json_data
    if not isinstance(enhanced, dict):
        return json_data
    return enhanced

def match_products_with_gemini(target_products: list, reference_products: list, prompt_template: str):
    fallback_result = {"matchedItems": [], "uniqueItems": target_products}
    if not target_products:
        return {"matchedItems": [], "uniqueItems": []}
    if not reference_products:
        return fallback_result
    match_prompt_formatted = prompt_template.format(
        target_products=json.dumps(target_products, ensure_ascii=False),
        reference_products=json.dumps(reference_products, ensure_ascii=False),
    )
    model = genai.GenerativeModel(
        model_name="gemini-2.5-pro",
        generation_config={"temperature": 0.1, "top_p": 0.95},
    )
    response = model.generate_content(match_prompt_formatted)
    match_text = (response.text or "").strip()
    match_data = extract_json_from_text(match_text)
    if not isinstance(match_data, dict) or "matchedItems" not in match_data or "uniqueItems" not in match_data:
        return fallback_result
    return match_data

def authenticate_and_open_sheet(sheet_id):
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=SCOPES)
    client = gspread.authorize(creds)
    return client.open_by_key(sheet_id).get_worksheet(0)

def ensure_first_three_rows_exist(ws):
    p = []
    for i in range(1, 4):
        p.append({"range": f"A{i}:B{i}", "values": [["", ""]]})
    ws.batch_update(p, value_input_option="USER_ENTERED")

def _last_non_empty_col_in_top_rows(ws):
    vals = ws.get_all_values()
    last = ITEM_MASTER_LIST_COL
    for row in vals[:HEADER_ROW]:
        for i, c in enumerate(row, start=1):
            if str(c).strip():
                last = max(last, i)
    return last

def find_next_available_column(ws):
    start_col = ITEM_MASTER_LIST_COL + 1
    last_used = _last_non_empty_col_in_top_rows(ws)
    if last_used < start_col:
        return start_col
    offset = last_used - start_col + 1
    groups_used = (offset + COLUMNS_PER_SUPPLIER - 1) // COLUMNS_PER_SUPPLIER
    return start_col + groups_used * COLUMNS_PER_SUPPLIER

def update_google_sheet_for_single_file(ws, data):
    ensure_first_three_rows_exist(ws)
    start_row = HEADER_ROW + 1
    sheet_values = ws.get_all_values()
    existing_products = []
    summary_row_map = {}
    first_summary_row = -1

    for row_idx, row in enumerate(sheet_values[HEADER_ROW:], start=start_row):
        if len(row) >= ITEM_MASTER_LIST_COL and row[ITEM_MASTER_LIST_COL - 1].strip():
            cell_value = row[ITEM_MASTER_LIST_COL - 1].strip()
            if cell_value in SUMMARY_LABELS:
                if first_summary_row == -1:
                    first_summary_row = row_idx
                summary_row_map[cell_value] = row_idx
            else:
                product_name = clean_product_name(cell_value)
                existing_products.append({"name": product_name, "row": row_idx})

    existing_suppliers = {}
    header_row_values = sheet_values[COMPANY_NAME_ROW - 1] if sheet_values else []
    for col_idx in range(
        ITEM_MASTER_LIST_COL + 1,
        len(header_row_values) + 1,
        COLUMNS_PER_SUPPLIER,
    ):
        supplier_name = ""
        if COMPANY_NAME_ROW - 1 < len(sheet_values) and (col_idx - 1) < len(sheet_values[COMPANY_NAME_ROW - 1]):
            supplier_name = sheet_values[COMPANY_NAME_ROW - 1][col_idx - 1].strip()
        if supplier_name:
            existing_suppliers[supplier_name] = col_idx

    next_avail_col = find_next_available_column(ws)

    products = data.get("products", [])
    if not products:
        return 0

    for product in products:
        if product.get("name"):
            product["name"] = clean_product_name(product["name"])

    company_name = data.get("company", "Unknown Company")
    col_idx = existing_suppliers.get(company_name, next_avail_col)
    if col_idx == next_avail_col:
        next_avail_col += COLUMNS_PER_SUPPLIER

    batch_requests = [
        {
            "range": f"{get_column_letter(col_idx)}{COMPANY_NAME_ROW}",
            "values": [[company_name]],
        },
        {
            "range": f"{get_column_letter(col_idx)}{CONTACT_INFO_ROW}",
            "values": [[f"{data.get('contact','')}".strip()]],
        },
        {
            "range": f"{get_column_letter(col_idx)}{HEADER_ROW}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{HEADER_ROW}",
            "values": [["ปริมาณ", "หน่วย", "ราคาต่อหน่วย", "รวมเป็นเงิน"]],
        },
    ]

    reference_data = [{"name": item["name"]} for item in existing_products]
    match_results = match_products_with_gemini(products, reference_data, matching_prompt)
    matched_items = match_results["matchedItems"]
    unique_items = match_results["uniqueItems"]

    populated_rows = set()
    for item in matched_items:
        item_name = item.get("name", "")
        for existing in existing_products:
            if existing["name"] == item_name and existing["row"] not in populated_rows:
                batch_requests.append(
                    {
                        "range": f"{get_column_letter(col_idx)}{existing['row']}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{existing['row']}",
                        "values": [[
                            item.get("quantity", 1),
                            item.get("unit", "ชิ้น"),
                            item.get("pricePerUnit", 0),
                            item.get("totalPrice", 0),
                        ]],
                    }
                )
                populated_rows.add(existing["row"])
                break

    new_products = []
    for item in unique_items:
        if isinstance(item, dict) and "name" in item:
            item["name"] = clean_product_name(item["name"])
            if not any(existing["name"] == item["name"] for existing in existing_products):
                new_products.append(item)

    insertion_row = first_summary_row if first_summary_row > 0 else (start_row + len(existing_products))

    if new_products:
        new_rows = [[""] * ws.col_count for _ in range(len(new_products))]
        if new_rows:
            ws.insert_rows(new_rows, insertion_row)

        row_shift = len(new_products)
        if first_summary_row > 0:
            for label in list(summary_row_map.keys()):
                summary_row_map[label] += row_shift

        for i, product in enumerate(new_products):
            row = insertion_row + i
            batch_requests.append(
                {
                    "range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}",
                    "values": [[product.get("name", "Unknown Product")]],
                }
            )
            batch_requests.append(
                {
                    "range": f"{get_column_letter(col_idx)}{row}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{row}",
                    "values": [[
                        product.get("quantity", 1),
                        product.get("unit", "ชิ้น"),
                        product.get("pricePerUnit", 0),
                        product.get("totalPrice", 0),
                    ]],
                }
            )

    price_col = col_idx + COLUMNS_PER_SUPPLIER - 1
    summary_items = [
        ("รวมเป็นเงิน", data.get("totalPrice", 0)),
        ("ภาษีมูลค่าเพิ่ม 7%", data.get("totalVat", 0)),
        ("ยอดรวมทั้งสิ้น", data.get("totalPriceIncludeVat", 0)),
        ("กำหนดยืนราคา (วัน)", data.get("priceGuaranteeDay", "")),
        ("ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO", data.get("deliveryTime", "")),
        ("การชำระเงิน", data.get("paymentTerms", "")),
        ("อื่น ๆ", data.get("otherNotes", "")),
    ]

    if summary_row_map:
        for label, value in summary_items:
            if label in summary_row_map:
                batch_requests.append(
                    {
                        "range": f"{get_column_letter(price_col)}{summary_row_map[label]}",
                        "values": [[value]],
                    }
                )
    else:
        summary_row = insertion_row + len(new_products) + 2
        for i, (label, value) in enumerate(summary_items):
            row = summary_row + i
            batch_requests.append(
                {
                    "range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}",
                    "values": [[label]],
                }
            )
            batch_requests.append(
                {"range": f"{get_column_letter(price_col)}{row}", "values": [[value]]}
            )

    if batch_requests:
        ws.batch_update(batch_requests, value_input_option="USER_ENTERED")

    return 1

def update_google_sheet_with_multiple_files(ws, all_json_data):
    if len(all_json_data) == 1:
        return update_google_sheet_for_single_file(ws, all_json_data[0])

    ensure_first_three_rows_exist(ws)
    start_row = HEADER_ROW + 1
    sheet_values = ws.get_all_values()
    existing_products = []

    for row_idx, row in enumerate(sheet_values[HEADER_ROW:], start=start_row):
        if len(row) >= ITEM_MASTER_LIST_COL and row[ITEM_MASTER_LIST_COL - 1].strip():
            cell_value = row[ITEM_MASTER_LIST_COL - 1].strip()
            if cell_value not in SUMMARY_LABELS:
                product_name = clean_product_name(cell_value)
                existing_products.append({"name": product_name, "row": row_idx})

    existing_suppliers = {}
    header_row_values = sheet_values[COMPANY_NAME_ROW - 1] if sheet_values else []
    for col_idx in range(
        ITEM_MASTER_LIST_COL + 1,
        len(header_row_values) + 1,
        COLUMNS_PER_SUPPLIER,
    ):
        supplier_name = ""
        if (col_idx - 1) < len(header_row_values):
            supplier_name = header_row_values[col_idx - 1].strip()
        if supplier_name:
            existing_suppliers[supplier_name] = col_idx

    next_avail_col = find_next_available_column(ws)

    for data in all_json_data:
        products = data.get("products", [])
        if not products:
            continue

        for product in products:
            if product.get("name"):
                product["name"] = clean_product_name(product["name"])

        company_name = data.get("company", "Unknown Company")
        col_idx = existing_suppliers.get(company_name, next_avail_col)
        if col_idx == next_avail_col:
            next_avail_col += COLUMNS_PER_SUPPLIER

        batch_requests = [
            {
                "range": f"{get_column_letter(col_idx)}{COMPANY_NAME_ROW}",
                "values": [[company_name]],
            },
            {
                "range": f"{get_column_letter(col_idx)}{CONTACT_INFO_ROW}",
                "values": [[f"{data.get('contact','')}".strip()]],
            },
            {
                "range": f"{get_column_letter(col_idx)}{HEADER_ROW}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{HEADER_ROW}",
                "values": [["ปริมาณ", "หน่วย", "ราคาต่อหน่วย", "รวมเป็นเงิน"]],
            },
        ]

        reference_data = [{"name": item["name"]} for item in existing_products]
        match_results = match_products_with_gemini(products, reference_data, matching_prompt)
        matched_items = match_results["matchedItems"]
        unique_items = match_results["uniqueItems"]

        populated_rows = set()

        for item in matched_items:
            item_name = item.get("name", "")
            for existing in existing_products:
                if existing["name"] == item_name and existing["row"] not in populated_rows:
                    batch_requests.append(
                        {
                            "range": f"{get_column_letter(col_idx)}{existing['row']}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{existing['row']}",
                            "values": [[
                                item.get("quantity", 1),
                                item.get("unit", "ชิ้น"),
                                item.get("pricePerUnit", 0),
                                item.get("totalPrice", 0),
                            ]],
                        }
                    )
                    populated_rows.add(existing["row"])
                    break

        new_products = []
        for item in unique_items:
            if isinstance(item, dict) and "name" in item:
                item["name"] = clean_product_name(item["name"])
                if not any(existing["name"] == item["name"] for existing in existing_products):
                    new_products.append(item)

        next_row = start_row + len(existing_products)

        if new_products:
            new_rows = [[""] * ws.col_count for _ in range(len(new_products))]
            if new_rows:
                ws.insert_rows(new_rows, next_row)

            for i, product in enumerate(new_products):
                row = next_row + i
                batch_requests.append(
                    {
                        "range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}",
                        "values": [[product.get("name", "Unknown Product")]],
                    }
                )
                batch_requests.append(
                    {
                        "range": f"{get_column_letter(col_idx)}{row}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{row}",
                        "values": [[
                            product.get("quantity", 1),
                            product.get("unit", "ชิ้น"),
                            product.get("pricePerUnit", 0),
                            product.get("totalPrice", 0),
                        ]],
                    }
                )
                existing_products.append({"name": product.get("name", "Unknown Product"), "row": row})

        summary_row = start_row + len(existing_products) + 2
        summary_items = [
            ("รวมเป็นเงิน", data.get("totalPrice", 0)),
            ("ภาษีมูลค่าเพิ่ม 7%", data.get("totalVat", 0)),
            ("ยอดรวมทั้งสิ้น", data.get("totalPriceIncludeVat", 0)),
            ("กำหนดยืนราคา (วัน)", data.get("priceGuaranteeDay", "")),
            ("ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO", data.get("deliveryTime", "")),
            ("การชำระเงิน", data.get("paymentTerms", "")),
            ("อื่น ๆ", data.get("otherNotes", "")),
        ]

        for i, (label, value) in enumerate(summary_items):
            row = summary_row + i
            batch_requests.append(
                {"range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}", "values": [[label]]}
            )
            batch_requests.append(
                {"range": f"{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{row}", "values": [[value]]}
            )

        if batch_requests:
            ws.batch_update(batch_requests, value_input_option="USER_ENTERED")

        if company_name not in existing_suppliers:
            existing_suppliers[company_name] = col_idx

    return len(all_json_data)

def get_file_type(file_path):
    mime_type, _ = mimetypes.guess_type(file_path)
    if mime_type:
        if mime_type.startswith("image/"):
            return "image"
        elif mime_type == "application/pdf":
            return "pdf"
        elif mime_type in [
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ]:
            return "word"
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".gif", ".webp"]:
        return "image"
    elif ext == ".pdf":
        return "pdf"
    elif ext in [".doc", ".docx"]:
        return "word"
    return "unknown"

def _wait_for_file_active(uploaded_file, timeout=180, poll=1.0):
    start = time.time()
    name = getattr(uploaded_file, "name", None)
    if not name:
        return uploaded_file
    while time.time() - start < timeout:
        try:
            f2 = genai.get_file(name)
            state = getattr(f2, "state", None)
            if state == "ACTIVE":
                return f2
            time.sleep(poll)
        except Exception:
            time.sleep(poll)
    return uploaded_file

def process_file(file_path):
    original_file_name = os.path.basename(file_path)
    safe_display_name = sanitize_filename_for_upload(original_file_name)
    with open(file_path, "rb") as src:
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_file_name)[1]) as tmp_file:
            tmp_file.write(src.read())
            tmp_file_path = tmp_file.name
    uploaded_gemini_file = genai.upload_file(path=tmp_file_path, display_name=safe_display_name)
    uploaded_gemini_file = _wait_for_file_active(uploaded_gemini_file)
    file_type = get_file_type(file_path)
    prompt_to_use = image_prompt if file_type == "image" else prompt
    model_flash = genai.GenerativeModel(model_name="gemini-2.5-flash", generation_config={"temperature": 0.1, "top_p": 0.95})
    resp = model_flash.generate_content([prompt_to_use, uploaded_gemini_file])
    d = extract_json_from_text(getattr(resp, "text", "") or "")
    if not d or not d.get("products"):
        st.info(f"Retrying with Pro model for {original_file_name}...")
        model_pro = genai.GenerativeModel(model_name="gemini-2.5-pro", generation_config={"temperature": 0.1, "top_p": 0.95})
        resp_pro = model_pro.generate_content([prompt_to_use, uploaded_gemini_file])
        d = extract_json_from_text(getattr(resp_pro, "text", "") or "")
    d = validate_json_data(d) if d else None
    d = enhance_with_gemini(d, validation_prompt) if d else None
    if d:
        result = {"file_name": original_file_name, "data": d}
    else:
        result = {"file_name": original_file_name, "error": "Failed to extract structured data from the document after multiple attempts."}
    if tmp_file_path and os.path.exists(tmp_file_path):
        os.unlink(tmp_file_path)
    if uploaded_gemini_file and getattr(uploaded_gemini_file, "name", None):
        genai.delete_file(uploaded_gemini_file.name)
    return result

def process_files(file_paths, sheet_id=DEFAULT_SHEET_ID):
    data_list = []
    error_list = []
    total_files = len(file_paths)
    with st.status(f"Processing {total_files} files...", expanded=True) as status:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(total_files, 10)) as executor:
            future_to_file = {executor.submit(process_file, path): path for path in file_paths}
            processed_count = 0
            for future in concurrent.futures.as_completed(future_to_file):
                processed_count += 1
                result = future.result()
                if "error" in result and result["error"]:
                    st.warning(f"⚠️ Failed to process {result['file_name']}: {result['error']}")
                    error_list.append(result)
                else:
                    st.write(f"✓ Successfully processed {result['file_name']}")
                    processed_data = result.get("data")
                    if isinstance(processed_data, list):
                        data_list.extend(processed_data)
                    elif isinstance(processed_data, dict):
                        data_list.append(processed_data)
                status.progress(processed_count / total_files, text=f"Processed {processed_count}/{total_files} files")
    if data_list:
        status.update(label="Updating Google Sheet...", state="running")
        ws = authenticate_and_open_sheet(sheet_id)
        update_google_sheet_with_multiple_files(ws, data_list)
        status.update(label="Processing complete!", state="complete", expanded=False)
    elif not error_list:
        status.update(label="No data could be extracted from the files.", state="complete")
    else:
        status.update(label="Processing finished with errors.", state="error", expanded=True)
    return data_list, error_list

def process_pdfs(pdf_paths, sheet_id=DEFAULT_SHEET_ID):
    return process_files(pdf_paths, sheet_id)

prompt = """# System Message for Intelligent Document Extraction
## Core Task: Single Document, Multiple Formats
Your primary task is to analyze a single uploaded document. This document can be one of two types:
1.  **A standard quotation** from a single company.
2.  **A multi-supplier comparison sheet**, which contains data from several companies in one view (e.g., side-by-side columns for 'Hisense', 'Franke', 'TEKA').

You MUST correctly identify the format and adapt your output accordingly.

## Output Specification
- For a **standard quotation**, you MUST return a **single JSON object**.
- For a **multi-supplier comparison sheet**, you MUST return a JSON **array** (a list of JSON objects). Each object in the array MUST correspond to one supplier.

## How to Identify and Process a Comparison Sheet
- **Detection:** Look for a layout with a central product description column (often the leftmost column with text) and multiple, repeating groups of columns for different company names or brands.
- **Processing Logic:** For each supplier you identify:
    1. Create a complete, separate JSON object for that supplier.
    2. Extract the supplier's `company` name and `contact` info from their specific header section.
    3. For each product row, you MUST associate the product description from the **central/master list column** with that specific supplier's columns for `quantity`, `unit`, `pricePerUnit`, etc.
    4. Calculate totals (`totalPrice`, `totalVat`, `totalPriceIncludeVat`) based on that supplier's data ONLY.
- **CRITICAL:** Do NOT merge data between suppliers into a single object. Each supplier gets their own object within the main list.

---
## Data Schema: Single JSON Object Format (for standard quotations)
{
  "company": "company name or first name + last name (NEVER null)",
  "vat": true,
  "contact": "phone number or email or null",
  "priceGuaranteeDay": 30,
  "deliveryTime": "",
  "paymentTerms": "",
  "otherNotes": "",
  "products": [
    {
      "name": "full product description including ALL parent category info",
      "quantity": 1,
      "unit": "match the unit shown in the document",
      "pricePerUnit": 0,
      "totalPrice": 0
    }
  ],
  "totalPrice": 0,
  "totalVat": 0,
  "totalPriceIncludeVat": 0
}

## Data Schema: JSON Array Format (for comparison sheets)
[
  {
    "company": "Supplier A Name",
    "contact": "Supplier A Contact",
    "products": [
        { "name": "Master Product 1 Name", "quantity": 10, "pricePerUnit": 100, "totalPrice": 1000 },
        { "name": "Master Product 2 Name", "quantity": 20, "pricePerUnit": 200, "totalPrice": 4000 }
    ],
    "totalPrice": 5000,
    "totalVat": 350,
    "totalPriceIncludeVat": 5350
  },
  {
    "company": "Supplier B Name",
    "contact": "Supplier B Contact",
    "products": [
        { "name": "Master Product 1 Name", "quantity": 10, "pricePerUnit": 105, "totalPrice": 1050 },
        { "name": "Master Product 2 Name", "quantity": 20, "pricePerUnit": 210, "totalPrice": 4200 }
    ],
    "totalPrice": 5250,
    "totalVat": 367.5,
    "totalPriceIncludeVat": 5617.5
  }
]
---

## General Extraction Rules
- **Accuracy:** You MUST ONLY extract information that is EXPLICITLY visible in the document. DO NOT HALLUCINATE.
- **Completeness:** Extract EVERY product line item for EACH supplier you identify.
- **Product Names:** When processing a comparison sheet, the `name` for a product should come from the shared, master product list column. For a standard quotation, include all hierarchical information.
- **Leading Numbers:** REMOVE any leading numbers (e.g., "1. ", "2) ") from product descriptions.
- **Prices & Quantities:** All numeric values (quantity, prices, totals) must be extracted precisely as numbers, without currency symbols or commas. Preserve full decimal precision.
- **Additional Details:** Extract `priceGuaranteeDay`, `deliveryTime`, `paymentTerms`, and `otherNotes` for each supplier if they are listed separately. If these details are shared, you may include them in each JSON object.
- **Final Verification:** Before concluding, review your generated JSON. Ensure the number of objects matches the number of suppliers and the number of products within each object matches the line items for that supplier.
"""

image_prompt = """# System Message for Intelligent Document Extraction
## Core Task: Single Document, Multiple Formats
Your primary task is to analyze a single uploaded document. This document can be one of two types:
1.  **A standard quotation** from a single company.
2.  **A multi-supplier comparison sheet**, which contains data from several companies in one view (e.g., side-by-side columns for 'Hisense', 'Franke', 'TEKA').

You MUST correctly identify the format and adapt your output accordingly.

## Output Specification
- For a **standard quotation**, you MUST return a **single JSON object**.
- For a **multi-supplier comparison sheet**, you MUST return a JSON **array** (a list of JSON objects). Each object in the array MUST correspond to one supplier.

## How to Identify and Process a Comparison Sheet
- **Detection:** Look for a layout with a central product description column (often the leftmost column with text) and multiple, repeating groups of columns for different company names or brands.
- **Processing Logic:** For each supplier you identify:
    1. Create a complete, separate JSON object for that supplier.
    2. Extract the supplier's `company` name and `contact` info from their specific header section.
    3. For each product row, you MUST associate the product description from the **central/master list column** with that specific supplier's columns for `quantity`, `unit`, `pricePerUnit`, etc.
    4. Calculate totals (`totalPrice`, `totalVat`, `totalPriceIncludeVat`) based on that supplier's data ONLY.
- **CRITICAL:** Do NOT merge data between suppliers into a single object. Each supplier gets their own object within the main list.

---
## Data Schema: Single JSON Object Format (for standard quotations)
{
  "company": "company name or first name + last name (NEVER null)",
  "vat": true,
  "contact": "phone number or email or null",
  "priceGuaranteeDay": 30,
  "deliveryTime": "",
  "paymentTerms": "",
  "otherNotes": "",
  "products": [
    {
      "name": "full product description including ALL parent category info",
      "quantity": 1,
      "unit": "match the unit shown in the document",
      "pricePerUnit": 0,
      "totalPrice": 0
    }
  ],
  "totalPrice": 0,
  "totalVat": 0,
  "totalPriceIncludeVat": 0
}

## Data Schema: JSON Array Format (for comparison sheets)
[
  {
    "company": "Supplier A Name",
    "contact": "Supplier A Contact",
    "products": [
        { "name": "Master Product 1 Name", "quantity": 10, "pricePerUnit": 100, "totalPrice": 1000 },
        { "name": "Master Product 2 Name", "quantity": 20, "pricePerUnit": 200, "totalPrice": 4000 }
    ],
    "totalPrice": 5000,
    "totalVat": 350,
    "totalPriceIncludeVat": 5350
  },
  {
    "company": "Supplier B Name",
    "contact": "Supplier B Contact",
    "products": [
        { "name": "Master Product 1 Name", "quantity": 10, "pricePerUnit": 105, "totalPrice": 1050 },
        { "name": "Master Product 2 Name", "quantity": 20, "pricePerUnit": 210, "totalPrice": 4200 }
    ],
    "totalPrice": 5250,
    "totalVat": 367.5,
    "totalPriceIncludeVat": 5617.5
  }
]
---

## General Extraction Rules
- **Accuracy:** You MUST ONLY extract information that is EXPLICITLY visible in the document. DO NOT HALLUCINATE.
- **Completeness:** Extract EVERY product line item for EACH supplier you identify.
- **Product Names:** When processing a comparison sheet, the `name` for a product should come from the shared, master product list column. For a standard quotation, include all hierarchical information.
- **Leading Numbers:** REMOVE any leading numbers (e.g., "1. ", "2) ") from product descriptions.
- **Prices & Quantities:** All numeric values (quantity, prices, totals) must be extracted precisely as numbers, without currency symbols or commas. Preserve full decimal precision.
- **Additional Details:** Extract `priceGuaranteeDay`, `deliveryTime`, `paymentTerms`, and `otherNotes` for each supplier if they are listed separately. If these details are shared, you may include them in each JSON object.
- **Final Verification:** Before concluding, review your generated JSON. Ensure the number of objects matches the number of suppliers and the number of products within each object matches the line items for that supplier.
"""

validation_prompt = """
You are a data validation expert specializing in Thai construction quotations.
I've extracted product data from a document, but there may be missing products or hierarchical relationships.

## CRITICAL: COMPLETE DATA CHECK
Your primary task is to ensure ALL products are correctly extracted with their hierarchical structure:
1. Check that all products visible in the document have been extracted
2. Ensure parent-child relationships and category groupings are preserved
3. Verify that all products have complete descriptions including their category name
4. Make sure no products are missing dimensions or specifications

## CRITICAL: PRESERVE PRODUCT HIERARCHY
Thai construction quotations often organize products hierarchically by categories:
- Category names with descriptive details
- Materials and specifications
- Dimensions

Each product must include its complete hierarchy:
"[Category Name] - [Material] - [Type] - [Dimensions]"

Examples (DO NOT add data from the attached Example in the prompt): 
- "งานบันไดกระจก งานพื้นตก - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม."
- "งานพื้นตก (ชั้นลอย) - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - เทมเปอร์ใส หนา 10 มม. ขนาด 3.565×0.97 ม."

## CRITICAL: DECIMAL NUMBER ACCURACY
Pay special attention to:
1. Quantities with decimals (extract full precision)
2. Dimensions with decimals (preserve exact measurements)
3. Prices with decimals (maintain exact values)

## CRITICAL: CLEAN PRODUCT DESCRIPTIONS
1. REMOVE any leading numbers (1., 2., 3., etc.) from product descriptions
2. Ensure NO product descriptions begin with numbering
3. Maintain all other hierarchical information and details

Review the data carefully and FIX these issues:
1. ADD any missing products that should be extracted from the source document
2. FIX product names to include complete hierarchical information WITHOUT leading numbers
3. ENSURE all dimensions and specifications are preserved with full decimal precision
4. VERIFY every product has the correct quantity, unit, price and total with full decimal precision

Original extraction:
{extracted_json}

Return ONLY a valid JSON object with no explanations.
"""

matching_prompt = """
You are a meticulous data architect specializing in product ontology for construction and home appliance materials. Your primary mission is to analyze product lists from different suppliers, establish a single "canonical" master product name for each item, and then map all supplier variations to that canonical name.
Your logic must be hierarchical and rule-based. Follow this algorithm precisely.

### **Core Objective: Create and Match to a Canonical Name**

The "Canonical Name" is the single source of truth for a product. You must construct it using this strict format:
**`[Group] - [Normalized Product Type] - [Primary Model/Identifier]`**

-   **`[Group]`**: The project phase or room size (e.g., "1BR+2BR(57-70sqm.)", "2BR (90sqm.)"). This is the **highest-priority** matching key.
-   **`[Normalized Product Type]`**: The generic category of the product (e.g., "Hood", "Induction Hob", "Sink"). You must deduce this from various descriptions.
-   **`[Primary Model/Identifier]`**: The most specific model number available (e.g., "EL 60", "MWE 255 FI").

### **Mandatory 4-Step Matching Algorithm**

For every product you process, you must follow these steps in order:

#### **Step 1: Group Matching (Non-Negotiable Filter)**
-   This is the most critical step. Products can **ONLY** be considered a match if they belong to the **exact same `[Group]`**.
-   Example: A "Hood" from "1BR+2BR(57-70sqm.)" can **NEVER** match a "Hood" from "2BR (90sqm.)". They are distinct line items.
-   Recognize semantic equivalents for groups, e.g., "1 BEDROOM" is the same as "1BR+2BR(57-70sqm.)".

#### **Step 2: Product Type Normalization & Keyword Mapping**
-   After filtering by group, identify the core product type. You must normalize different supplier descriptions into one standard type.
-   Use this keyword map as your guide:
    -   **"Hood"**: `Slimline Hood`, `BI telescopic hood`, `HOOD PIAVE 60 XS`
    -   **"Induction Hob"**: `Induction Hob`, `Hob Electric`, `HOB INDUCTION`
    -   **"Microwave"**: `Built-in Microwave`, `Microwave Oven`, `MICROWAVE FMWO 25 NH I`
    -   **"Sink"**: `Undermount Sink`, `Sink Stainless Steel`, `SINK BXX 210-45`
    -   **"Tap"**: `Sink Single Tap`, `Tap`, `TAP LANNAR`

#### **Step 3: Specification & Model Analysis**
-   Once Group and Normalized Type match, use the model number and other specifications (`Model`, `Description` columns) to create the full canonical name and to confirm the match.
-   The model number itself does not have to be identical between suppliers if the Group and Normalized Type are a clear match. The model number's purpose is to create the *unique canonical name* for that row.

#### **Step 4: Construct Final Output**
-   Based on the matches found, generate the final JSON.

### **Critical Rules & Constraints**

1.  **Group is King:** If the group doesn't match, nothing else matters.
2.  **Type over Model:** A strong match on `Group` + `Normalized Product Type` is more important than a weak match on `Model` number.
3.  **One-to-One Mapping:** A reference item (a canonical name you create) can only be matched once per supplier list.
4.  **No Imagination:** Only use information explicitly present in the data. If you cannot confidently normalize a product type, classify it as unique.

### **Walkthrough Example: Matching "Hoods"**

**Goal:** Match the first item from all three suppliers.

1.  **Teka:**
    -   **Input:** Group=`1BR+2BR(57-70sqm.)`, Model=`EL 60`, Desc=`Slimline Hood`
    -   **Analysis:** Group is "1BR...". Type normalizes from "Slimline Hood" to **"Hood"**. Model is `EL 60`.
    -   **Canonical Name Created:** `1BR+2BR(57-70sqm.) - Hood - EL 60`

2.  **Hisense (Gorenje):**
    -   **Input:** Group=`1BR+2BR (57-70Sqm.)`, Product=`TH62E3X`, Desc=`BI telescopic hood...`
    -   **Analysis:** Group is "1BR...". It matches Teka's group. Type normalizes from "BI telescopic hood" to **"Hood"**. It matches the normalized type.
    -   **Conclusion:** This is a match for the same row.

3.  **Franke:**
    -   **Input:** Group=`1 BEDROOM`, Product Category=`Hood`, Mode=`PIAVE 60 XS`
    -   **Analysis:** Group "1 BEDROOM" is semantically identical to "1BR...". It matches. Type is explicitly `Hood`. It matches.
    -   **Conclusion:** This is also a match for the same row.

All three products are mapped to the canonical name `1BR+2BR(57-70sqm.) - Hood - EL 60`, and their respective data will be aligned on this single row in the final output.

### **Input & Output Format**

-   **Input:** `target_products` (from a new quotation) and `reference_products` (the existing master list of canonical names).
-   **Output:** You **MUST** return a JSON object with this exact structure:

{{
  "matchedItems": [
    {{
      "name": "The canonical reference name this product matched to.",
      "quantity": "target quantity",
      "unit": "target unit",
      "pricePerUnit": "target price per unit",
      "totalPrice": "target total price"
    }}
  ],
  "uniqueItems": [
    {{
      "name": "The full, descriptive name of the target product that could not be matched.",
      "quantity": "target quantity",
      "unit": "target unit",
      "pricePerUnit": "target price per unit",
      "totalPrice": "target total price"
    }}
  ]
}}

## Target Products:
{target_products}

## Reference Products:
{reference_products}
"""

def main():
    st.set_page_config(page_title="ระบบประมวลผลใบเสนอราคา", layout="centered")
    st.sidebar.title("ตั้งค่าการเชื่อมต่อ")
    google_api_key = st.sidebar.text_input(
        "Enter your GOOGLE_API_KEY", 
        value=st.session_state.get("google_api_key", ""), 
        type="password",
        key="google_api_key_input"
    )

    if st.sidebar.button("ยืนยัน API Key", key="confirm_api_key", use_container_width=True):
        if google_api_key:
            genai.configure(api_key=google_api_key)
            st.session_state.google_api_key = google_api_key
            st.session_state.api_key_confirmed = True
            st.sidebar.success("API Key ถูกบันทึกและใช้งานได้แล้ว")
        else:
            st.session_state.api_key_confirmed = False
            st.sidebar.warning("กรุณาใส่ API Key ก่อนยืนยัน")

    st.markdown("<h1 style='text-align: center;'>ระบบประมวลผลใบเสนอราคา</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>อัปโหลดไฟล์ใบเสนอราคา (PDF หรือรูปภาพ) เพื่อดึงข้อมูลและบันทึกลง Google Sheet โดยอัตโนมัติ</p>", unsafe_allow_html=True)
    st.markdown("---")
    st.subheader("ขั้นตอนที่ 1: อัปโหลดข้อมูลของคุณ")
    sheet_url = st.text_input(
        "Google Sheet URL or ID:",
        value=DEFAULT_SHEET_ID,
        placeholder="ใส่ URL หรือ ID ของ Google Sheet ที่ต้องการบันทึกข้อมูล"
    )
    uploaded_files = st.file_uploader(
        "เลือกไฟล์ PDF หรือรูปภาพ (สามารถเลือกได้หลายไฟล์พร้อมกัน)",
        type=['pdf', 'jpg', 'jpeg', 'png'],
        accept_multiple_files=True
    )
    if uploaded_files:
        if st.button("🚀 เริ่มประมวลผล", use_container_width=True, type="primary"):
            if not st.session_state.get("api_key_confirmed"):
                st.error("❌ กรุณาใส่และยืนยัน Google API Key ในแถบด้านข้างก่อนเริ่มประมวลผล")
                return
            file_paths = []
            with st.spinner("กำลังเตรียมไฟล์เพื่ออัปโหลด..."):
                for uploaded_file in uploaded_files:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[1]) as tmp_file:
                        tmp_file.write(uploaded_file.getvalue())
                        file_paths.append(tmp_file.name)
            sheet_id = extract_sheet_id_from_url(sheet_url) if sheet_url else DEFAULT_SHEET_ID
            results, errors = process_files(file_paths, sheet_id)
            st.subheader("ขั้นตอนที่ 2: ผลการประมวลผล")
            if results:
                st.success(f"✅ ประมวลผลสำเร็จ {len(results)} ไฟล์ และบันทึกข้อมูลลง Google Sheet เรียบร้อยแล้ว")
                sheet_url_display = f"https://docs.google.com/spreadsheets/d/{sheet_id}"
                st.markdown(f"**[คลิกที่นี่เพื่อเปิด Google Sheet]({sheet_url_display})**")
                for result_data in results:
                    company_name = result_data.get('company', 'Unknown Company')
                    with st.expander(f"📄 ผลลัพธ์จาก: {company_name}"):
                        st.json(result_data)
            if errors:
                st.error(f"❌ พบข้อผิดพลาด {len(errors)} ไฟล์ ไม่สามารถประมวลผลได้")
                for error_info in errors:
                    with st.expander(f"🚨 ข้อผิดพลาดในไฟล์: {error_info['file_name']}"):
                        st.write(f"**สาเหตุ:**")
                        st.code(error_info['error'], language=None)
            if not results and not errors:
                st.warning("ไม่สามารถประมวลผลไฟล์ใดๆ ได้ กรุณาตรวจสอบไฟล์และลองใหม่อีกครั้ง")

if __name__ == "__main__":
    main()