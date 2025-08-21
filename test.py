import concurrent.futures
import os
import json
import tempfile
import re
import time
import mimetypes

import google.generativeai as genai
import gspread
import streamlit as st

from google.oauth2.service_account import Credentials
from openpyxl.utils import get_column_letter

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

DEFAULT_SHEET_ID = "17tMHStXQYXaIQHQIA4jdUyHaYt_tuoNCEEuJCstWEuw"

COMPANY_NAME_ROW = 1
CONTACT_INFO_ROW = 2
HEADER_ROW = 3
ITEM_MASTER_LIST_COL = 2
COLUMNS_PER_SUPPLIER = 4

SUMMARY_LABELS = [
    "รวมเป็นเงิน",
    "ภาษีมูลค่าเพิ่ม 7%",
    "ยอดรวมทั้งสิ้น",
    "กำหนดยืนราคา (วัน)",
    "ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO",
    "การชำระเงิน",
    "อื่น ๆ",
]

SAFETY_SETTINGS = {
    "HARM_CATEGORY_HARASSMENT": "BLOCK_NONE",
    "HARM_CATEGORY_HATE_SPEECH": "BLOCK_NONE",
    "HARM_CATEGORY_SEXUALLY_EXPLICIT": "BLOCK_NONE",
    "HARM_CATEGORY_DANGEROUS_CONTENT": "BLOCK_NONE",
}

def extract_sheet_id_from_url(url):
    if not url:
        return None
    if "/" not in url and " " not in url and len(url) > 20:
        return url
    m = re.search(r"spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None

def extract_json_from_text(text):
    if not text:
        return None
    blocks = re.findall(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", text)
    if blocks:
        candidates = blocks
    else:
        start = text.find("{")
        end = text.rfind("}") + 1
        candidates = [text[start:end]] if start >= 0 and end > start else []

    for cand in candidates:
        json_str = cand
        cleaned_json = re.sub(r",\s*}", "}", json_str)
        cleaned_json = re.sub(r",\s*]", "]", cleaned_json)
        try:
            return json.loads(cleaned_json)
        except json.JSONDecodeError:
            continue
    return None

def extract_contact_info(text):
    if not text:
        return ""
    phone_pattern = r"(?<!\w)((0\d{1,2}[-\s]?\d{3}[-\s]?\d{3,4})|(0\d{2}[-\s]?\d{7})|(0\d{2}[-\s]?\d{3}[-\s]?\d{4}))(?!\w)"
    email_pattern = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
    phone_matches = re.findall(phone_pattern, text)
    email_matches = re.findall(email_pattern, text)
    phone_numbers = [m[0] for m in phone_matches] if phone_matches else []
    emails = sorted(set(email_matches))
    phones = []
    for phone in phone_numbers:
        clean_phone = re.sub(r"\s", "", phone)
        if len(clean_phone) >= 9:
            phones.append(clean_phone)
    phones = sorted(set(phones))
    contact_parts = []
    if emails:
        contact_parts.append(f"Email: {', '.join(emails)}")
    if phones:
        contact_parts.append(f"Phone: {', '.join(phones)}")
    return ", ".join(contact_parts)

def clean_product_name(name):
    if not name:
        return "Unknown Product"
    return re.sub(r"^\s*\d+[\.\)\-]\s*", "", name.strip())

def _to_number_or_default(val, default):
    s = str(val)
    s2 = s.replace(",", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", s2):
        return float(s2)
    return default

def validate_json_data(json_data):
    if not json_data:
        return {
            "company": "Unknown Company",
            "contact": "",
            "vat": False,
            "products": [],
            "totalPrice": 0,
            "totalVat": 0,
            "totalPriceIncludeVat": 0,
            "priceGuaranteeDay": 0,
            "deliveryTime": "",
            "paymentTerms": "",
            "otherNotes": "",
        }
    if not json_data.get("company"):
        json_data["company"] = "Unknown Company"

    if "contact" in json_data:
        if isinstance(json_data["contact"], dict):
            contact_parts = []
            if "email" in json_data["contact"] and json_data["contact"]["email"]:
                contact_parts.append(f"Email: {json_data['contact']['email']}")
            if "phone" in json_data["contact"] and json_data["contact"]["phone"]:
                contact_parts.append(f"Phone: {json_data['contact']['phone']}")
            json_data["contact"] = ", ".join(contact_parts)
        else:
            json_data["contact"] = extract_contact_info(str(json_data["contact"]))
    else:
        json_data["contact"] = ""

    if "vat" not in json_data:
        json_data["vat"] = False
    else:
        json_data["vat"] = bool(json_data["vat"])

    if not json_data.get("products"):
        json_data["products"] = []
    for product in json_data.get("products", []):
        product["name"] = clean_product_name(product.get("name") or "Unknown Product")
        product["quantity"] = _to_number_or_default(product.get("quantity", 1), 1)
        if product["quantity"] <= 0:
            product["quantity"] = 1
        product["unit"] = product.get("unit") or "ชิ้น"
        product["pricePerUnit"] = _to_number_or_default(product.get("pricePerUnit", 0), 0)
        provided_total = _to_number_or_default(product.get("totalPrice", None), None)
        if provided_total is None:
            product["totalPrice"] = round(product["quantity"] * product["pricePerUnit"], 2)
        else:
            product["totalPrice"] = provided_total

    computed_total = sum(p.get("totalPrice", 0) for p in json_data.get("products", []))
    json_data["totalPrice"] = _to_number_or_default(json_data.get("totalPrice", computed_total), computed_total)
    json_data["totalVat"] = _to_number_or_default(json_data.get("totalVat", round(json_data["totalPrice"] * 0.07, 2)), round(json_data["totalPrice"] * 0.07, 2))
    json_data["totalPriceIncludeVat"] = _to_number_or_default(
        json_data.get("totalPriceIncludeVat", round(json_data["totalPrice"] + json_data["totalVat"], 2)),
        round(json_data["totalPrice"] + json_data["totalVat"], 2),
    )

    if "priceGuaranteeDay" not in json_data:
        json_data["priceGuaranteeDay"] = 0
    if "deliveryTime" not in json_data:
        json_data["deliveryTime"] = ""
    if "paymentTerms" not in json_data:
        json_data["paymentTerms"] = ""
    if "otherNotes" not in json_data:
        json_data["otherNotes"] = ""

    return json_data

def match_products_with_gemini(target_products, reference_products):
    if not target_products:
        return {"matchedItems": [], "uniqueItems": []}
    if not reference_products:
        return {"matchedItems": [], "uniqueItems": target_products}

    try:
        match_prompt_formatted = matching_prompt.format(
            target_products=json.dumps(target_products, ensure_ascii=False),
            reference_products=json.dumps(reference_products, ensure_ascii=False),
        )
    except KeyError as e:
        match_prompt_formatted = matching_prompt.replace("{matchedItems}", "").replace("{uniqueItems}", "")
        match_prompt_formatted = match_prompt_formatted.format(
            target_products=json.dumps(target_products, ensure_ascii=False),
            reference_products=json.dumps(reference_products, ensure_ascii=False),
        )

    model = genai.GenerativeModel(
        model_name="gemini-2.5-pro",
        generation_config={"temperature": 0.0, "top_p": 0.95},
        safety_settings=SAFETY_SETTINGS
    )
    response = model.generate_content(match_prompt_formatted)
    match_text = (response.text or "").strip()
    match_data = extract_json_from_text(match_text)

    if not match_data:
        start = match_text.find("{")
        end = match_text.rfind("}") + 1
        if start >= 0 and end > start:
            try:
                match_data = json.loads(match_text[start:end])
            except Exception:
                match_data = None

    if not match_data or not isinstance(match_data, dict):
        return {"matchedItems": [], "uniqueItems": target_products}
    if "matchedItems" not in match_data or not isinstance(match_data["matchedItems"], list):
        match_data["matchedItems"] = []
    if "uniqueItems" not in match_data or not isinstance(match_data["uniqueItems"], list):
        match_data["uniqueItems"] = target_products
    return match_data

def authenticate_and_open_sheet(sheet_id):
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=SCOPES)
    client = gspread.authorize(creds)
    return client.open_by_key(sheet_id).get_worksheet(0)

def ensure_first_three_rows_exist(ws):
    p = []
    for i in range(1, 4):
        p.append({"range": f"A{i}:B{i}", "values": [["", ""]]})
    ws.batch_update(p, value_input_option="USER_ENTERED")

def _last_non_empty_col_in_top_rows(ws):
    vals = ws.get_all_values()
    last = ITEM_MASTER_LIST_COL
    for row in vals[:HEADER_ROW]:
        for i, c in enumerate(row, start=1):
            if str(c).strip():
                last = max(last, i)
    return last

def find_next_available_column(ws):
    start_col = ITEM_MASTER_LIST_COL + 1
    last_used = _last_non_empty_col_in_top_rows(ws)
    if last_used < start_col:
        return start_col
    offset = last_used - start_col + 1
    groups_used = (offset + COLUMNS_PER_SUPPLIER - 1) // COLUMNS_PER_SUPPLIER
    return start_col + groups_used * COLUMNS_PER_SUPPLIER

def update_google_sheet_for_single_file(ws, data):
    ensure_first_three_rows_exist(ws)
    start_row = HEADER_ROW + 1
    sheet_values = ws.get_all_values()
    existing_products = []
    summary_row_map = {}
    first_summary_row = -1

    for row_idx, row in enumerate(sheet_values[HEADER_ROW:], start=start_row):
        if len(row) >= ITEM_MASTER_LIST_COL and row[ITEM_MASTER_LIST_COL - 1].strip():
            cell_value = row[ITEM_MASTER_LIST_COL - 1].strip()
            if cell_value in SUMMARY_LABELS:
                if first_summary_row == -1:
                    first_summary_row = row_idx
                summary_row_map[cell_value] = row_idx
            else:
                product_name = clean_product_name(cell_value)
                existing_products.append({"name": product_name, "row": row_idx})

    existing_suppliers = {}
    header_row_values = sheet_values[COMPANY_NAME_ROW - 1] if sheet_values else []
    for col_idx in range(
        ITEM_MASTER_LIST_COL + 1,
        len(header_row_values) + 1,
        COLUMNS_PER_SUPPLIER,
    ):
        supplier_name = ""
        if COMPANY_NAME_ROW - 1 < len(sheet_values) and (col_idx - 1) < len(sheet_values[COMPANY_NAME_ROW - 1]):
            supplier_name = sheet_values[COMPANY_NAME_ROW - 1][col_idx - 1].strip()
        if supplier_name:
            existing_suppliers[supplier_name] = col_idx

    next_avail_col = find_next_available_column(ws)

    products = data.get("products", [])
    if not products:
        return

    for product in products:
        if product.get("name"):
            product["name"] = clean_product_name(product["name"])

    company_name = data.get("company", "Unknown Company")
    col_idx = existing_suppliers.get(company_name, next_avail_col)

    batch_requests = [
        {"range": f"{get_column_letter(col_idx)}{COMPANY_NAME_ROW}", "values": [[company_name]]},
        {"range": f"{get_column_letter(col_idx)}{CONTACT_INFO_ROW}", "values": [[f"{data.get('contact', '')}".strip()]]},
        {"range": f"{get_column_letter(col_idx)}{HEADER_ROW}:{get_column_letter(col_idx + COLUMNS_PER_SUPPLIER - 1)}{HEADER_ROW}", "values": [["ปริมาณ", "หน่วย", "ราคาต่อหน่วย", "รวมเป็นเงิน"]]},
    ]

    reference_data = [{"name": item["name"]} for item in existing_products]
    match_results = match_products_with_gemini(products, reference_data)
    matched_items = match_results.get("matchedItems", [])
    unique_items = match_results.get("uniqueItems", [])

    populated_rows = set()
    for item in matched_items:
        item_name = item.get("name", "")
        for existing in existing_products:
            if existing["name"] == item_name and existing["row"] not in populated_rows:
                batch_requests.append({
                    "range": f"{get_column_letter(col_idx)}{existing['row']}:{get_column_letter(col_idx + COLUMNS_PER_SUPPLIER - 1)}{existing['row']}",
                    "values": [[item.get("quantity", 1), item.get("unit", "ชิ้น"), item.get("pricePerUnit", 0), item.get("totalPrice", 0)]],
                })
                populated_rows.add(existing["row"])
                break

    new_products = []
    for item in unique_items:
        if isinstance(item, dict) and "name" in item:
            item["name"] = clean_product_name(item["name"])
            if not any(existing["name"] == item["name"] for existing in existing_products):
                new_products.append(item)

    insertion_row = first_summary_row if first_summary_row > 0 else (start_row + len(existing_products))

    if new_products:
        ws.insert_rows([[""] * ws.col_count for _ in new_products], insertion_row)
        row_shift = len(new_products)
        if first_summary_row > 0:
            for label in list(summary_row_map.keys()):
                summary_row_map[label] += row_shift

        for i, product in enumerate(new_products):
            row = insertion_row + i
            batch_requests.append({"range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}", "values": [[product.get("name", "Unknown Product")]]})
            batch_requests.append({
                "range": f"{get_column_letter(col_idx)}{row}:{get_column_letter(col_idx + COLUMNS_PER_SUPPLIER - 1)}{row}",
                "values": [[product.get("quantity", 1), product.get("unit", "ชิ้น"), product.get("pricePerUnit", 0), product.get("totalPrice", 0)]],
            })

    price_col = col_idx + COLUMNS_PER_SUPPLIER - 1
    summary_items = [
        ("รวมเป็นเงิน", data.get("totalPrice", 0)),
        ("ภาษีมูลค่าเพิ่ม 7%", data.get("totalVat", 0)),
        ("ยอดรวมทั้งสิ้น", data.get("totalPriceIncludeVat", 0)),
        ("กำหนดยืนราคา (วัน)", data.get("priceGuaranteeDay", "")),
        ("ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO", data.get("deliveryTime", "")),
        ("การชำระเงิน", data.get("paymentTerms", "")),
        ("อื่น ๆ", data.get("otherNotes", "")),
    ]

    if summary_row_map:
        for label, value in summary_items:
            if label in summary_row_map:
                batch_requests.append({"range": f"{get_column_letter(price_col)}{summary_row_map[label]}", "values": [[value]]})
    else:
        summary_row = insertion_row + len(new_products) + 2
        for i, (label, value) in enumerate(summary_items):
            row = summary_row + i
            batch_requests.append({"range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}", "values": [[label]]})
            batch_requests.append({"range": f"{get_column_letter(price_col)}{row}", "values": [[value]]})

    if batch_requests:
        ws.batch_update(batch_requests, value_input_option="USER_ENTERED")

def get_file_type(file_path):
    mime_type, _ = mimetypes.guess_type(file_path)
    if mime_type:
        if mime_type.startswith("image/"):
            return "image"
        elif mime_type == "application/pdf":
            return "pdf"
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".jpg", ".jpeg", ".png"]:
        return "image"
    elif ext == ".pdf":
        return "pdf"
    return "unknown"

def _wait_for_file_active(uploaded_file, timeout=180, poll=1.0):
    start = time.time()
    name = getattr(uploaded_file, "name", None)
    if not name:
        return uploaded_file
    while time.time() - start < timeout:
        f2 = genai.get_file(name)
        state = getattr(f2, "state", None)
        if state == "ACTIVE":
            return f2
        time.sleep(poll)
    return uploaded_file

def process_file(file_path):
    file_name = os.path.basename(file_path)
    tmp_file_path = None
    uploaded_gemini_file = None
    
    with open(file_path, "rb") as src:
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file_name)[1]) as tmp_file:
            tmp_file.write(src.read())
            tmp_file_path = tmp_file.name

    uploaded_gemini_file = genai.upload_file(path=tmp_file_path, display_name=file_name)
    uploaded_gemini_file = _wait_for_file_active(uploaded_gemini_file)

    file_type = get_file_type(file_path)
    prompt_to_use = image_prompt if file_type == "image" else prompt

    model_flash = genai.GenerativeModel(model_name="gemini-2.5-flash", generation_config={"temperature": 0.1, "top_p": 0.95}, safety_settings=SAFETY_SETTINGS)
    resp = model_flash.generate_content([prompt_to_use, uploaded_gemini_file])
    d = extract_json_from_text(getattr(resp, "text", "") or "")

    if not d or not d.get("products"):
        model_pro = genai.GenerativeModel(model_name="gemini-2.5-pro", generation_config={"temperature": 0.1, "top_p": 0.95}, safety_settings=SAFETY_SETTINGS)
        resp_pro = model_pro.generate_content([prompt_to_use, uploaded_gemini_file])
        d = extract_json_from_text(getattr(resp_pro, "text", "") or "")

    d = validate_json_data(d) if d else None
    
    result = {"file_name": file_name, "data": d}
    
    if tmp_file_path:
        os.unlink(tmp_file_path)
    if uploaded_gemini_file:
        genai.delete_file(uploaded_gemini_file.name)
        
    return result

def process_files(file_paths, sheet_id=DEFAULT_SHEET_ID):
    data_by_index = {}
    total_files = len(file_paths)
    
    with st.status(f"Processing {total_files} files...", expanded=True) as status:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(total_files, 8)) as executor:
            future_to_index = {executor.submit(process_file, path): idx for idx, path in enumerate(file_paths)}
            processed_count = 0
            for future in concurrent.futures.as_completed(future_to_index):
                idx = future_to_index[future]
                processed_count += 1
                result = future.result()
                st.write(f"✓ Extracted {result['file_name']}")
                data_by_index[idx] = result
                status.progress(processed_count / total_files, text=f"Extracted {processed_count}/{total_files} files")

    results = []
    if data_by_index:
        status.update(label="Updating Google Sheet sequentially...", state="running")
        ws = authenticate_and_open_sheet(extract_sheet_id_from_url(sheet_id) or DEFAULT_SHEET_ID)
        
        sorted_indices = sorted(data_by_index.keys())
        
        for i, idx in enumerate(sorted_indices):
            r = data_by_index[idx]
            if r and "data" in r and r["data"]:
                st.write(f"🔁 Updating sheet with file {i+1}/{len(sorted_indices)}: {r['file_name']}")
                update_google_sheet_for_single_file(ws, r["data"])
                results.append(r["data"])
                time.sleep(0.5)
            
        status.update(label="Processing complete!", state="complete", expanded=False)
        
    return results, []

prompt = """# System Message for Product List Extraction (PDF/Text Table Processing)
## CRITICAL: ANTI-HALLUCINATION WARNING
You MUST ONLY extract information that is EXPLICITLY visible in the document. 
- DO NOT add data from anywhere other than the one in the uploaded document. Adding data that is not from the document is a serious mistake.
- DO NOT create or invent any products, prices, or specifications
- DO NOT add data from the attached Example in the prompt.
- DO NOT add products that are not clearly listed as distinct line items
- DO NOT attempt to break down a single product into multiple products
- DO NOT interpret descriptive text as separate products
- If uncertain about any information, LEAVE IT OUT rather than guessing

## CRITICAL: CONTACT INFORMATION EXTRACTION
Pay special attention to contact information in the document header or footer:
- Extract any email addresses (example@domain.com)
- Extract any phone numbers (formats like 02-3384825, 081-1234567, 095-525-2623)
- Put email and phone details in the "contact" field
- Format: "Email: email@example.com, Phone: 081-234-5678"
- Use only the phone number and email address on the letterhead. Do not add "บริษัท ลูก้า แอสเซท จำกัด, 081-781-7283" contact information

## CRITICAL: COMPLETE EXTRACTION REQUIREMENT
You MUST extract ALL products visible in the document:
- Extract EVERY product line item visible in the document
- Preserve hierarchical structure (groups/categories) of products if present
- Ensure NO products are missed or skipped
- Each row with a distinct price is one product

## Input Format
Provide PDF files (or images/tables with text extraction) containing product information (receipts, invoices, product lists, etc.).

## Task
Extract ALL product information EXPLICITLY visible in the document:
- Extract products with quantities, units, and prices
- Preserve parent-child relationships (main categories and sub-items)
- Maintain hierarchical product groupings (numbered sections, categories)
- Also extract additional quotation details like price validity, delivery time, payment terms, etc.

## Hierarchical Structure Handling
Many quotations organize products hierarchically. When you see this:
- Include category names in product descriptions (e.g., "งานบันไดกระจก งานพื้นตก - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม.")
- If product descriptions begin with numbers (1, 2, 3...), REMOVE those numbers
- Include all parent category information in each product's name without the leading numbers

## Output Format (JSON only)
You must return ONLY this JSON structure:
{
  "company": "company name or first name + last name (NEVER null)",
  "vat": true,
  "name": "customer name or null",
  "contact": "phone number or email or null",
  "priceGuaranteeDay": 30
  "deliveryTime": "",
  "paymentTerms": "",
  "otherNotes": "",
  "products": [
    {
      "name": "full product description including ALL parent category info, specifications AND dimensions WITHOUT leading numbers",
      "quantity": 1,
      "unit": "match the unit shown in the document (e.g., แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, ชุด)",
      "pricePerUnit": 0,
      "totalPrice": 0
    }
  ],
  "totalPrice": 0,
  "totalVat": 0,
  "totalPriceIncludeVat": 0
}

## Example 1 (Format with Item/ART.No./Description/Qty/Unit/Price columns):
| Item | ART.No. | Description | Qty | Unit | Standard Price | Discount Price | Amount |
|------|---------|-------------|-----|------|----------------|---------------|--------|
| 1    | CPW-xxxx| SPC ลายไม้ 4.5 มิล (ก้างปลา) | 1.00 | ตร.ม. |  | 520.00 | 520.000 |
| 2    |         | ค่าแรงติดตั้ง | 1.00 | ตร.ม. |  | 150.00 | 150.000 |

## Example 2 (Format with ลำดับ/รหัสสินค้า/รายละเอียดสินค้า columns):
| ลำดับ | รหัสสินค้า | รายละเอียดสินค้า | หน่วย | จำนวน | ราคา/หน่วย(บาท) | จำนวนเงิน(บาท) |
|------|---------|----------------|------|------|--------------|------------|
| 1    |         | พื้นไม้ไวนิลลายไม้ปลา 4.5 มม. LKT 4.5 mm x 0.3 mm สีฟ้าเซอร์คูลี (1 กล่อง บรรจุ 18 แผ่น หรือ 1.3 ตร.ม) | ตร.ม. | 1.30 | 680.00 | 884.00 |

## Field Extraction Guidelines

### name (Product Description)
* CRITICAL: Include ALL hierarchical information in each product name:
  - Category names/headings (e.g., "งานบันไดกระจก งานพื้นตก")
  - Sub-category information (e.g., "เหล็กตัวซีชุบสังกะสี")
  - Glass type, thickness (e.g., "กระจกเทมเปอร์ใส หนา 10 มม.")
  - Exact dimensions (e.g., "ขนาด 4.672×0.97 ม.")
* REMOVE any leading numbers (1., 2., 3.) from the product descriptions
* Format hierarchical products as: "[Category Name] - [Material] - [Type] - [Dimensions]"
* Include: ALL distinguishing characteristics that make each product unique
* Example: "งานบันไดกระจก งานพื้นตก - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม."

### unit and quantity (DIRECT EXTRACTION RULE)
* Extract unit and quantity DIRECTLY from each line item as shown
* Use the exact unit shown in the document (ชุด, แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, etc.)
* Extract the exact quantity shown for each product (never assume or calculate)
* NEVER create quantities or units that aren't explicitly shown in the document
* Pay special attention to decimal quantities - extract the full decimal precision

### pricePerUnit and totalPrice
* Extract ONLY prices clearly visible in the document
* Use numeric values only (no currency symbols)
* Extract cleanly from pricing fields as shown in each line item
* NEVER calculate or estimate prices that aren't explicitly shown
* NEVER combine different products' prices
* Pay special attention to decimal prices - extract the EXACT decimal values shown

### Additional Quotation Details
* Extract these additional fields if present:
  - "กำหนดยืนราคา (วัน)", "กำหนดยืนราคา", "การยืนราคา" - Price validity period in days (priceGuaranteeDay)
  - "ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO" - Delivery time after PO (deliveryTime)
  - "การชำระเงิน" - Payment terms (paymentTerms)
  - "อื่น ๆ" - Other notes (otherNotes)
* Extract as text exactly as written, preserving numbers and Thai language

### CRITICAL: Pricing summaries and summary values
* Extract the exact values for these three summary items:
  - "รวมเป็นเงิน" - the initial subtotal (totalPrice)
  - "ภาษีมูลค่าเพิ่ม 7%" - the VAT amount (totalVat)
  - "ยอดรวมทั้งสิ้น" - the final total (totalPriceIncludeVat)
* Alternative labels to match:
  - For totalPrice: "รวม", "รวมเป็นเงิน", "ราคารวม", "Total", "TOTAL AMOUNT", "รวมราคา"
  - For totalVat: "ภาษีมูลค่าเพิ่ม 7%", "VAT 7%"
  - For totalPriceIncludeVat: "ยอดรวมทั้งสิ้น", "รวมทั้งหมด", "รวมเงินทั้งสิน", "ราคารวมสุทธิ", "รวมราคางานทั้งหมดตามสัญญา"
* Extract the exact values as shown (remove commas, currency symbols)
* CRITICAL: Preserve full decimal precision in all monetary values

## FINAL VERIFICATION
Review the extracted products one last time and verify:
1. Count the number of products you've extracted
2. Verify this matches EXACTLY with the number of product rows visible in the document
3. Check that ALL products have proper hierarchical information included WITHOUT leading numbers
4. Ensure NO products are missing - every line item with a price must be extracted
5. Confirm all dimensions and specifications are preserved correctly
6. Verify all decimal values (quantities and prices) maintain their full precision
"""

image_prompt = """# System Message for Product List Extraction from Images
## CRITICAL: ANTI-HALLUCINATION WARNING
You MUST ONLY extract information that is EXPLICITLY visible in the image. 
- DO NOT create or invent any products, prices, or specifications
- DO NOT add products that are not clearly listed as distinct line items
- DO NOT interpret descriptive text as separate products
- If text is unclear or unreadable, mark it as uncertain rather than guessing

## CRITICAL: COMPLETE EXTRACTION REQUIREMENT
You MUST extract ALL products visible in the image:
- Extract EVERY product line item visible in the image
- Preserve hierarchical structure (groups/categories) of products if present
- Ensure NO products are missed or skipped
- Each row with a distinct price is one product

## Input Format
I'm providing an image of a document containing product information.

## Task
Extract ALL product information EXPLICITLY visible in the image:
- Extract products with quantities, units, and prices
- Preserve parent-child relationships (main categories and sub-items)
- Maintain hierarchical product groupings (numbered sections, categories)
- Also extract additional quotation details like price validity, delivery time, payment terms, etc.

## Hierarchical Structure Handling
Many quotations organize products hierarchically. When you see this:
- Include category names in product descriptions (e.g., "งานบันไดกระจก งานพื้นตก - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม.")
- If product descriptions begin with numbers (1, 2, 3...), REMOVE those numbers
- Include all parent category information in each product's name without the leading numbers

## Output Format (JSON only)
You must return ONLY this JSON structure:
{
  "company": "company name or first name + last name (NEVER null)",
  "vat": true,
  "contact": "phone number or email or null",
  "priceGuaranteeDay": 30
  "deliveryTime": "",
  "paymentTerms": "",
  "otherNotes": "",
  "products": [
    {
      "name": "full product description including ALL parent category info, specifications AND dimensions WITHOUT leading numbers",
      "quantity": 1,
      "unit": "match the unit shown in the document (e.g., แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, ชุด)",
      "pricePerUnit": 0,
      "totalPrice": 0
    }
  ],
  "totalPrice": 0,
  "totalVat": 0,
  "totalPriceIncludeVat": 0
}


## Example 1 (Format with Item/ART.No./Description/Qty/Unit/Price columns):
| Item | ART.No. | Description | Qty | Unit | Standard Price | Discount Price | Amount |
|------|---------|-------------|-----|------|----------------|---------------|--------|
| 1    | CPW-xxxx| SPC ลายไม้ 4.5 มิล (ก้างปลา) | 1.00 | ตร.ม. |  | 520.00 | 520.000 |
| 2    |         | ค่าแรงติดตั้ง | 1.00 | ตร.ม. |  | 150.00 | 150.000 |

## Example 2 (Format with ลำดับ/รหัสสินค้า/รายละเอียดสินค้า columns):
| ลำดับ | รหัสสินค้า | รายละเอียดสินค้า | หน่วย | จำนวน | ราคา/หน่วย(บาท) | จำนวนเงิน(บาท) |
|------|---------|----------------|------|------|--------------|------------|
| 1    |         | พื้นไม้ไวนิลลายไม้ปลา 4.5 มม. LKT 4.5 mm x 0.3 mm สีฟ้าเซอร์คูลี (1 กล่อง บรรจุ 18 แผ่น หรือ 1.3 ตร.ม) | ตร.ม. | 1.30 | 680.00 | 884.00 |


## Field Extraction Guidelines

### name (Product Description)
* CRITICAL: Include ALL hierarchical information in each product name:
  - Category names/headings (e.g., "งานบันไดกระจก งานพื้นตก")
  - Sub-category information (e.g., "เหล็กตัวซีชุบสังกะสี")
  - Glass type, thickness (e.g., "กระจกเทมเปอร์ใส หนา 10 มม.")
  - Exact dimensions (e.g., "ขนาด 4.672×0.97 ม.")
* REMOVE any leading numbers (1., 2., 3.) from the product descriptions
* Format hierarchical products as: "[Category Name] - [Material] - [Type] - [Dimensions]"
* Include: ALL distinguishing characteristics that make each product unique
* Example: "งานบันไดกระจก งานพื้นตก - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม."

### unit and quantity (DIRECT EXTRACTION RULE)
* Extract unit and quantity DIRECTLY from each line item as shown
* Use the exact unit shown in the document (ชุด, แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, จำนวนต่อชุด etc.)
* Extract the exact quantity shown for each product (never assume or calculate)
* NEVER create quantities or units that aren't explicitly shown in the document
* Pay special attention to decimal quantities - extract the full decimal precision

### pricePerUnit and totalPrice
* Extract ONLY prices clearly visible in the document
* Use numeric values only (no currency symbols)
* Extract cleanly from pricing fields as shown in each line item
* NEVER calculate or estimate prices that aren't explicitly shown
* NEVER combine different products' prices
* Pay special attention to decimal prices - extract the EXACT decimal values shown

### Additional Quotation Details
* Extract these additional fields if present:
  - "กำหนดยืนราคา (วัน)", "กำหนดยืนราคา", "การยืนราคา" - Price validity period in days (priceGuaranteeDay)
  - "ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO" - Delivery time after PO (deliveryTime)
  - "การชำระเงิน" - Payment terms (paymentTerms)
  - "อื่น ๆ" - Other notes (otherNotes)
* Extract as text exactly as written, preserving numbers and Thai language

### CRITICAL: Pricing summaries and summary values
* Extract the exact values for these three summary items:
  - "รวมเป็นเงิน" - the initial subtotal (totalPrice)
  - "ภาษีมูลค่าเพิ่ม 7%" - the VAT amount (totalVat)
  - "ยอดรวมทั้งสิ้น" - the final total (totalPriceIncludeVat)
* Alternative labels to match:
  - For totalPrice: "รวม", "รวมเป็นเงิน", "ราคารวม", "Total", "TOTAL AMOUNT", "รวมราคา"
  - For totalVat: "ภาษีมูลค่าเพิ่ม 7%", "VAT 7%"
  - For totalPriceIncludeVat: "ยอดรวมทั้งสิ้น", "รวมทั้งหมด", "รวมเงินทั้งสิน", "ราคารวมสุทธิ", "รวมราคางานทั้งหมดตามสัญญา"
* Extract the exact values as shown (remove commas, currency symbols)
* CRITICAL: Preserve full decimal precision in all monetary values

## FINAL VERIFICATION
Review the extracted products one last time and verify:
1. Count the number of products you've extracted
2. Verify this matches EXACTLY with the number of product rows visible in the image
3. Check that ALL products have proper hierarchical information included WITHOUT leading numbers
4. Ensure NO products are missing - every line item with a price must be extracted
5. Confirm all dimensions and specifications are preserved correctly
6. Verify all decimal values (quantities and prices) maintain their full precision
"""

validation_prompt = """
You are a data validation expert specializing in Thai construction quotations.
I've extracted product data from a document, but there may be missing products or hierarchical relationships.

## CRITICAL: COMPLETE DATA CHECK
Your primary task is to ensure ALL products are correctly extracted with their hierarchical structure:
1. Check that all products visible in the document have been extracted
2. Ensure parent-child relationships and category groupings are preserved
3. Verify that all products have complete descriptions including their category name
4. Make sure no products are missing dimensions or specifications

## CRITICAL: PRESERVE PRODUCT HIERARCHY
Thai construction quotations often organize products hierarchically by categories:
- Category names with descriptive details
- Materials and specifications
- Dimensions

Each product must include its complete hierarchy:
"[Category Name] - [Material] - [Type] - [Dimensions]"

Examples (DO NOT add data from the attached Example in the prompt): 
- "งานบันไดกระจก งานพื้นตก - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม."
- "งานพื้นตก (ชั้นลอย) - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - เทมเปอร์ใส หนา 10 มม. ขนาด 3.565×0.97 ม."

## CRITICAL: DECIMAL NUMBER ACCURACY
Pay special attention to:
1. Quantities with decimals (extract full precision)
2. Dimensions with decimals (preserve exact measurements)
3. Prices with decimals (maintain exact values)

## CRITICAL: CLEAN PRODUCT DESCRIPTIONS
1. REMOVE any leading numbers (1., 2., 3., etc.) from product descriptions
2. Ensure NO product descriptions begin with numbering
3. Maintain all other hierarchical information and details

Review the data carefully and FIX these issues:
1. ADD any missing products that should be extracted from the source document
2. FIX product names to include complete hierarchical information WITHOUT leading numbers
3. ENSURE all dimensions and specifications are preserved with full decimal precision
4. VERIFY every product has the correct quantity, unit, price and total with full decimal precision

Original extraction:
{extracted_json}

Return ONLY a valid JSON object with no explanations.
"""

matching_prompt = """
You are a meticulous data architect specializing in product ontology for construction and home appliance materials. Your primary mission is to analyze product lists from different suppliers, establish a single "canonical" master product name for each item, and then map all supplier variations to that canonical name.
Your logic must be hierarchical and rule-based. Follow this algorithm precisely.

### **Core Objective: Create and Match to a Canonical Name**

The "Canonical Name" is the single source of truth for a product. You must construct it using this strict format:
**`[Group] - [Normalized Product Type] - [Primary Model/Identifier]`**

-   **`[Group]`**: The project phase or room size (e.g., "1BR+2BR(57-70sqm.)", "2BR (90sqm.)"). This is the **highest-priority** matching key.
-   **`[Normalized Product Type]`**: The generic category of the product (e.g., "Hood", "Induction Hob", "Sink"). You must deduce this from various descriptions.
-   **`[Primary Model/Identifier]`**: The most specific model number available (e.g., "EL 60", "MWE 255 FI").

### **Mandatory 4-Step Matching Algorithm**

For every product you process, you must follow these steps in order:

#### **Step 1: Group Matching (Non-Negotiable Filter)**
-   This is the most critical step. Products can **ONLY** be considered a match if they belong to the **exact same `[Group]`**.
-   Example: A "Hood" from "1BR+2BR(57-70sqm.)" can **NEVER** match a "Hood" from "2BR (90sqm.)". They are distinct line items.
-   Recognize semantic equivalents for groups, e.g., "1 BEDROOM" is the same as "1BR+2BR(57-70sqm.)".

#### **Step 2: Product Type Normalization & Keyword Mapping**
-   After filtering by group, identify the core product type. You must normalize different supplier descriptions into one standard type.
-   Use this keyword map as your guide:
    -   **"Hood"**: `Slimline Hood`, `BI telescopic hood`, `HOOD PIAVE 60 XS`
    -   **"Induction Hob"**: `Induction Hob`, `Hob Electric`, `HOB INDUCTION`
    -   **"Microwave"**: `Built-in Microwave`, `Microwave Oven`, `MICROWAVE FMWO 25 NH I`
    -   **"Sink"**: `Undermount Sink`, `Sink Stainless Steel`, `SINK BXX 210-45`
    -   **"Tap"**: `Sink Single Tap`, `Tap`, `TAP LANNAR`

#### **Step 3: Specification & Model Analysis**
-   Once Group and Normalized Type match, use the model number and other specifications (`Model`, `Description` columns) to create the full canonical name and to confirm the match.
-   The model number itself does not have to be identical between suppliers if the Group and Normalized Type are a clear match. The model number's purpose is to create the *unique canonical name* for that row.

#### **Step 4: Construct Final Output**
-   Based on the matches found, generate the final JSON.

### **Critical Rules & Constraints**

1.  **Group is King:** If the group doesn't match, nothing else matters.
2.  **Type over Model:** A strong match on `Group` + `Normalized Product Type` is more important than a weak match on `Model` number.
3.  **One-to-One Mapping:** A reference item (a canonical name you create) can only be matched once per supplier list.
4.  **No Imagination:** Only use information explicitly present in the data. If you cannot confidently normalize a product type, classify it as unique.

### **Walkthrough Example: Matching "Hoods"**

**Goal:** Match the first item from all three suppliers.

1.  **Teka:**
    -   **Input:** Group=`1BR+2BR(57-70sqm.)`, Model=`EL 60`, Desc=`Slimline Hood`
    -   **Analysis:** Group is "1BR...". Type normalizes from "Slimline Hood" to **"Hood"**. Model is `EL 60`.
    -   **Canonical Name Created:** `1BR+2BR(57-70sqm.) - Hood - EL 60`

2.  **Hisense (Gorenje):**
    -   **Input:** Group=`1BR+2BR (57-70Sqm.)`, Product=`TH62E3X`, Desc=`BI telescopic hood...`
    -   **Analysis:** Group is "1BR...". It matches Teka's group. Type normalizes from "BI telescopic hood" to **"Hood"**. It matches the normalized type.
    -   **Conclusion:** This is a match for the same row.

3.  **Franke:**
    -   **Input:** Group=`1 BEDROOM`, Product Category=`Hood`, Mode=`PIAVE 60 XS`
    -   **Analysis:** Group "1 BEDROOM" is semantically identical to "1BR...". It matches. Type is explicitly `Hood`. It matches.
    -   **Conclusion:** This is also a match for the same row.

All three products are mapped to the canonical name `1BR+2BR(57-70sqm.) - Hood - EL 60`, and their respective data will be aligned on this single row in the final output.

### **Input & Output Format**

-   **Input:** `target_products` (from a new quotation) and `reference_products` (the existing master list of canonical names).
-   **Output:** You **MUST** return a JSON object with this exact structure:

{{
  "matchedItems": [
    {{
      "name": "The canonical reference name this product matched to.",
      "quantity": "target quantity",
      "unit": "target unit",
      "pricePerUnit": "target price per unit",
      "totalPrice": "target total price"
    }}
  ],
  "uniqueItems": [
    {{
      "name": "The full, descriptive name of the target product that could not be matched.",
      "quantity": "target quantity",
      "unit": "target unit",
      "pricePerUnit": "target price per unit",
      "totalPrice": "target total price"
    }}
  ]
}}

## Target Products:
{target_products}

## Reference Products:
{reference_products}
"""

def main():
    st.set_page_config(page_title="ระบบประมวลผลใบเสนอราคา", layout="centered")
    st.sidebar.title("ตั้งค่าการเชื่อมต่อ")
    google_api_key = st.sidebar.text_input(
        "Enter your GOOGLE_API_KEY", 
        value=st.session_state.get("google_api_key", "AIzaSyBRZwmkhzwW_XwFNxmH4CO_Ps62pYuxgI0"), 
        type="password",
        key="google_api_key_input"
    )

    if st.sidebar.button("ยืนยัน API Key", key="confirm_api_key", use_container_width=True):
        if google_api_key:
            genai.configure(api_key=google_api_key)
            st.session_state.google_api_key = google_api_key
            st.session_state.api_key_confirmed = True
            st.sidebar.success("API Key ถูกบันทึกและใช้งานได้แล้ว")
        else:
            st.session_state.api_key_confirmed = False
            st.sidebar.warning("กรุณาใส่ API Key ก่อนยืนยัน")

    st.markdown("<h1 style='text-align: center;'>ระบบประมวลผลใบเสนอราคา</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>อัปโหลดไฟล์ใบเสนอราคา (PDF หรือรูปภาพ) เพื่อดึงข้อมูลและบันทึกลง Google Sheet โดยอัตโนมัติ</p>", unsafe_allow_html=True)
    st.markdown("---")
    st.subheader("ขั้นตอนที่ 1: อัปโหลดข้อมูลของคุณ")
    sheet_url = st.text_input(
        "Google Sheet URL or ID:",
        value=DEFAULT_SHEET_ID,
        placeholder="ใส่ URL หรือ ID ของ Google Sheet ที่ต้องการบันทึกข้อมูล"
    )
    uploaded_files = st.file_uploader(
        "เลือกไฟล์ PDF หรือรูปภาพ (สามารถเลือกได้หลายไฟล์พร้อมกัน)",
        type=['pdf', 'jpg', 'jpeg', 'png'],
        accept_multiple_files=True
    )
    if uploaded_files:
        if st.button("🚀 เริ่มประมวลผล", use_container_width=True, type="primary"):
            if not st.session_state.get("api_key_confirmed"):
                st.error("❌ กรุณาใส่และยืนยัน Google API Key ในแถบด้านข้างก่อนเริ่มประมวลผล")
                return
                
            temp_dir = tempfile.mkdtemp()
            file_paths = []
            for uploaded_file in uploaded_files:
                file_path = os.path.join(temp_dir, uploaded_file.name)
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                file_paths.append(file_path)
            
            sheet_id = extract_sheet_id_from_url(sheet_url) if sheet_url else DEFAULT_SHEET_ID
            results, errors = process_files(file_paths, sheet_id)
            
            st.subheader("ขั้นตอนที่ 2: ผลการประมวลผล")
            if results:
                st.success(f"✅ ประมวลผลสำเร็จ {len(results)} ไฟล์ และบันทึกข้อมูลลง Google Sheet เรียบร้อยแล้ว")
                sheet_url_display = f"https://docs.google.com/spreadsheets/d/{sheet_id}"
                st.markdown(f"**[คลิกที่นี่เพื่อเปิด Google Sheet]({sheet_url_display})**")
                for result_data in results:
                    company_name = result_data.get('company', 'Unknown Company')
                    with st.expander(f"📄 ผลลัพธ์จาก: {company_name}"):
                        st.json(result_data)
            
            if not results and not errors:
                st.warning("ไม่สามารถประมวลผลไฟล์ใดๆ ได้ กรุณาตรวจสอบไฟล์และลองใหม่อีกครั้ง")

if __name__ == "__main__":
    main()