import concurrent.futures
import os
import json
import tempfile
import re
import time
import mimetypes

import google.generativeai as genai
import streamlit as st

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

# =========================================
# =========== CONFIG & CONSTANTS ==========
# =========================================

# บันทึกเป็นไฟล์ Excel (ถ้าไม่มีไฟล์จะสร้างใหม่อัตโนมัติ)
# ถ้ามีเทมเพลตอยู่แล้ว (เช่นที่อัปโหลดไว้) ใส่ path นี้ได้เลย
DEFAULT_EXCEL_PATH = "/mnt/data/test1.xlsx"

COMPANY_NAME_ROW = 1
CONTACT_INFO_ROW = 2
HEADER_ROW = 3
ITEM_MASTER_LIST_COL = 2
COLUMNS_PER_SUPPLIER = 4

SUMMARY_LABELS = [
    "รวมเป็นเงิน",
    "ภาษีมูลค่าเพิ่ม 7%",
    "ยอดรวมทั้งสิ้น",
    "กำหนดยืนราคา (วัน)",
    "ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO",
    "การชำระเงิน",
    "อื่น ๆ",
]

# =========================================
# =========== Excel Adapter Layer =========
# =========================================

class ExcelSheetAdapter:
    """
    ครอบ openpyxl ให้เมธอดคล้าย gspread พอประมาณ
    - get_all_values() -> list[list]
    - batch_update([{range, values}])
    - insert_rows(new_rows, insertion_row)
    - col_count property
    """
    def __init__(self, ws):
        self.ws = ws

    @property
    def col_count(self):
        return max(1, self.ws.max_column or 1)

    def get_all_values(self):
        max_row = max(HEADER_ROW + 1, self.ws.max_row or 0)
        max_col = max(ITEM_MASTER_LIST_COL + 4, self.ws.max_column or 0)
        data = []
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                v = self.ws.cell(row=r, column=c).value
                row_vals.append("" if v is None else str(v))
            # ตัดคอลัมน์ว่างท้ายแถว
            while row_vals and (row_vals[-1] == "" or row_vals[-1] is None):
                row_vals.pop()
            data.append(row_vals)
        # ตัดแถวว่างท้ายไฟล์
        while data and (len(data[-1]) == 0 or all(v == "" for v in data[-1])):
            data.pop()
        return data

    def batch_update(self, requests, value_input_option="USER_ENTERED"):
        for req in requests:
            rng = req["range"]
            values = req["values"]
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            r_cnt = max_row - min_row + 1
            c_cnt = max_col - min_col + 1
            for ridx in range(r_cnt):
                row_values = values[ridx] if ridx < len(values) else []
                for cidx in range(c_cnt):
                    val = row_values[cidx] if cidx < len(row_values) else ""
                    self.ws.cell(row=min_row + ridx, column=min_col + cidx, value=val)

    def insert_rows(self, new_rows, insertion_row):
        amount = len(new_rows)
        if amount > 0:
            self.ws.insert_rows(insertion_row, amount=amount)

# =========================================
# =========== Helpers (คงของเดิม) =========
# =========================================

def extract_sheet_id_from_url(url):
    if not url:
        return None
    if "/" not in url and " " not in url and len(url) > 20:
        return url
    m = re.search(r"spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None

def extract_json_from_text(text):
    if not text:
        return None
    blocks = re.findall(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", text)
    if blocks:
        candidates = blocks
    else:
        start = text.find("{")
        end = text.rfind("}") + 1
        candidates = [text[start:end]] if start >= 0 and end > start else []

    for cand in candidates:
        json_str = cand
        cleaned_json = re.sub(r",\s*}", "}", json_str)
        cleaned_json = re.sub(r",\s*]", "]", cleaned_json)
        try:
            return json.loads(cleaned_json)
        except json.JSONDecodeError:
            continue
    return None

def extract_contact_info(text):
    if not text:
        return ""
    phone_pattern = r"(?<!\w)((0\d{1,2}[-\s]?\d{3}[-\s]?\d{3,4})|(0\d{2}[-\s]?\d{7})|(0\d{2}[-\s]?\d{3}[-\s]?\d{4}))(?!\w)"
    email_pattern = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
    phone_matches = re.findall(phone_pattern, text)
    email_matches = re.findall(email_pattern, text)
    phone_numbers = [m[0] for m in phone_matches] if phone_matches else []
    emails = sorted(set(email_matches))
    phones = []
    for phone in phone_numbers:
        clean_phone = re.sub(r"\s", "", phone)
        if len(clean_phone) >= 9:
            phones.append(clean_phone)
    phones = sorted(set(phones))
    contact_parts = []
    if emails:
        contact_parts.append(f"Email: {', '.join(emails)}")
    if phones:
        contact_parts.append(f"Phone: {', '.join(phones)}")
    return ", ".join(contact_parts)

def clean_product_name(name):
    if not name:
        return "Unknown Product"
    return re.sub(r"^\s*\d+[\.\)\-]\s*", "", name.strip())

def _to_number_or_default(val, default):
    s = str(val)
    s2 = s.replace(",", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", s2):
        try:
            return float(s2)
        except Exception:
            return default
    return default

def validate_json_data(json_data):
    if not json_data:
        return {
            "company": "Unknown Company",
            "contact": "",
            "vat": False,
            "products": [],
            "totalPrice": 0,
            "totalVat": 0,
            "totalPriceIncludeVat": 0,
            "priceGuaranteeDay": 0,
            "deliveryTime": "",
            "paymentTerms": "",
            "otherNotes": "",
        }
    if not json_data.get("company"):
        json_data["company"] = "Unknown Company"

    if "contact" in json_data:
        if isinstance(json_data["contact"], dict):
            contact_parts = []
            if "email" in json_data["contact"] and json_data["contact"]["email"]:
                contact_parts.append(f"Email: {json_data['contact']['email']}")
            if "phone" in json_data["contact"] and json_data["contact"]["phone"]:
                contact_parts.append(f"Phone: {json_data['contact']['phone']}")
            json_data["contact"] = ", ".join(contact_parts)
        else:
            json_data["contact"] = extract_contact_info(str(json_data["contact"]))
    else:
        json_data["contact"] = ""

    if "vat" not in json_data:
        json_data["vat"] = False
    else:
        json_data["vat"] = bool(json_data["vat"])

    if not json_data.get("products"):
        json_data["products"] = []
    for product in json_data.get("products", []):
        product["name"] = clean_product_name(product.get("name") or "Unknown Product")
        product["quantity"] = _to_number_or_default(product.get("quantity", 1), 1)
        if product["quantity"] <= 0:
            product["quantity"] = 1
        product["unit"] = product.get("unit") or "ชิ้น"
        product["pricePerUnit"] = _to_number_or_default(product.get("pricePerUnit", 0), 0)
        provided_total = _to_number_or_default(product.get("totalPrice", None), None)
        if provided_total is None:
            product["totalPrice"] = round(product["quantity"] * product["pricePerUnit"], 2)
        else:
            product["totalPrice"] = provided_total

    computed_total = sum(p.get("totalPrice", 0) for p in json_data.get("products", []))
    json_data["totalPrice"] = _to_number_or_default(json_data.get("totalPrice", computed_total), computed_total)
    json_data["totalVat"] = _to_number_or_default(json_data.get("totalVat", round(json_data["totalPrice"] * 0.07, 2)), round(json_data["totalPrice"] * 0.07, 2))
    json_data["totalPriceIncludeVat"] = _to_number_or_default(
        json_data.get("totalPriceIncludeVat", round(json_data["totalPrice"] + json_data["totalVat"], 2)),
        round(json_data["totalPrice"] + json_data["totalVat"], 2),
    )

    if "priceGuaranteeDay" not in json_data:
        json_data["priceGuaranteeDay"] = 0
    if "deliveryTime" not in json_data:
        json_data["deliveryTime"] = ""
    if "paymentTerms" not in json_data:
        json_data["paymentTerms"] = ""
    if "otherNotes" not in json_data:
        json_data["otherNotes"] = ""

    return json_data

def enhance_with_gemini(json_data):
    if not json_data:
        return None
    validation_formatted = validation_prompt.format(
        extracted_json=json.dumps(json_data, ensure_ascii=False)
    )
    model = genai.GenerativeModel(model_name="gemini-2.5-flash")
    try:
        response = model.generate_content(validation_formatted)
        enhanced_text = (response.text or "").strip()
    except Exception:
        return json_data

    enhanced = extract_json_from_text(enhanced_text) or json_data
    if not isinstance(enhanced, dict):
        return json_data
    return enhanced

# =========================================
# =========== SMART MATCHING RULES ========
# =========================================

def _normalize_text(s: str) -> str:
    if not s:
        return ""
    s = s.lower().strip()
    s = s.replace("×", "x").replace("*", "x")
    s = re.sub(r"^\s*\d+[\.\)\-]\s*", "", s)
    s = re.sub(r"\s+", " ", s)
    return s

_UNIT_MM = {
    "มม": 1.0, "mm": 1.0,
    "ซม": 10.0, "cm": 10.0,
    "ม.": 1000.0, "เมตร": 1000.0, "m": 1000.0,
}

def _to_mm(val: float, unit: str) -> float:
    unit = (unit or "").lower().strip()
    if unit in _UNIT_MM:
        return val * _UNIT_MM[unit]
    if val <= 100:
        return val * 1000.0
    return val

_DIM_LABELS = {
    "กว้าง": "w", "width": "w", "w": "w",
    "สูง": "h", "height": "h", "h": "h",
    "ยาว": "l", "length": "l", "l": "l",
    "หนา": "t", "thick": "t", "thickness": "t", "t": "t",
    "ลึก": "d", "depth": "d", "d": "d",
}

def _parse_dimensions(text: str):
    res = {"labeled": {}, "sequence": []}
    s = _normalize_text(text)

    patt_labeled = r"(กว้าง|สูง|ยาว|หนา|ลึก|width|height|length|thick(?:ness)?|depth|[whltd])\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(มม|mm|ซม|cm|ม\.|เมตร|m)?"
    for lab, num, unit in re.findall(patt_labeled, s):
        key = _DIM_LABELS.get(lab, None)
        if key:
            mm = _to_mm(float(num), unit or "")
            res["labeled"][key] = mm

    patt_seq = r"(\d+(?:\.\d+)?)(?:\s*x\s*(\d+(?:\.\d+)?))(?:\s*x\s*(\d+(?:\.\d+)?))?\s*(มม|mm|ซม|cm|ม\.|เมตร|m)?"
    for a, b, c, unit in re.findall(patt_seq, s):
        nums = [float(a), float(b)]
        if c:
            nums.append(float(c))
        mm_vals = [_to_mm(v, unit or "") for v in nums]
        mm_vals.sort()
        if mm_vals not in res["sequence"]:
            res["sequence"].append(mm_vals)

    patt_thick = r"(?:หนา|thick(?:ness)?|t)?\s*(\d+(?:\.\d+)?)\s*(มม|mm)\b"
    for n, unit in re.findall(patt_thick, s):
        mm = _to_mm(float(n), unit)
        res["labeled"].setdefault("t", mm)

    return res

def _nearly_equal_mm(a: float, b: float, pct=0.02, abs_mm=5.0) -> bool:
    if a == 0 or b == 0:
        return abs(a - b) <= abs_mm
    return abs(a - b) <= max(abs_mm, pct * max(a, b))

def _dims_close(dim_a, dim_b) -> bool:
    la, lb = dim_a.get("labeled", {}), dim_b.get("labeled", {})
    inter = set(la.keys()) & set(lb.keys())
    if inter:
        ok = all(_nearly_equal_mm(la[k], lb[k]) for k in inter)
        if not ok:
            return False
    sa, sb = dim_a.get("sequence", []), dim_b.get("sequence", [])
    if sa and sb:
        if len(sa[0]) == len(sb[0]) and all(_nearly_equal_mm(x, y) for x, y in zip(sa[0], sb[0])):
            return True
        return False
    return True

_TYPE_KEYWORDS = {
    "aluminum_rail": ["รางอลูมิเนียม", "อลูมิเนียมโปรไฟล์", "aluminium rail", "aluminum rail", "ราง alu"],
    "steel_u": ["เหล็กตัวยู", "u-channel", "เหล็ก u", "รางยู", "เหล็กยู"],
    "glass_tempered": ["กระจกเทมเปอร์", "tempered glass", "เทมเปอร์"],
    "glass_laminate": ["กระจกลาามิเนต", "laminated glass", "laminate glass", "ลามิเนต"],
    "hinge": ["บานพับ", "hinge"],
    "handle": ["มือจับ", "handle", "knob", "pull"],
    "seal": ["ยางขอบ", "ซีล", "seal", "gasket"],
    "bracket": ["โช๊ค", "โช๊คอัพ", "bracket", "closer"],
    "frame": ["วงกบ", "กรอบ", "frame"],
}

def _normalize_types(text: str) -> set:
    s = _normalize_text(text)
    got = set()
    for tkey, kws in _TYPE_KEYWORDS.items():
        if any(kw in s for kw in kws):
            got.add(tkey)
    if ("กระจก" in s or "glass" in s) and not {"glass_tempered", "glass_laminate"} & got:
        got.add("glass_generic")
    return got

def _tokenize_material_semantics(text: str) -> set:
    s = _normalize_text(text)
    mats = set()
    if "อลูมิเนียม" in s or "aluminium" in s or "aluminum" in s:
        mats.add("aluminium")
    if "เหล็ก" in s or "steel" in s:
        mats.add("steel")
    if "กระจก" in s or "glass" in s:
        mats.add("glass")
    if "สแตนเลส" in s or "stainless" in s:
        mats.add("stainless")
    if "ตัวซี" in s or "c-channel" in s:
        mats.add("c_channel")
    if "ตัวยู" in s or "u-channel" in s:
        mats.add("u_channel")
    if "โปรไฟล์" in s:
        mats.add("profile")
    return mats

def _semantic_compatible(name_a: str, name_b: str) -> bool:
    ta, tb = _normalize_types(name_a), _normalize_types(name_b)
    if ta and tb and (ta & tb):
        return True
    ma, mb = _tokenize_material_semantics(name_a), _tokenize_material_semantics(name_b)
    return bool(ma & mb)

def match_products_smart(target_products, reference_products):
    """
    target_products: [{name, quantity, unit, pricePerUnit, totalPrice}, ...]
    reference_products: [{name}, ...]  ← baseline (มาจากไฟล์แรกที่บันทึกลง Excel แล้ว)
    """
    matched = []
    uniques = []

    ref_norm_map = {}
    ref_dims_map = {}
    for r in reference_products:
        rn = _normalize_text(r.get("name", ""))
        ref_norm_map[rn] = r.get("name", "")
        ref_dims_map[rn] = _parse_dimensions(r.get("name", ""))

    for item in target_products or []:
        nm = str(item.get("name", "")).strip()
        nm_clean = _normalize_text(nm)
        dims_t = _parse_dimensions(nm)

        # exact หลัง normalize
        if nm_clean in ref_norm_map:
            matched.append({
                "name": ref_norm_map[nm_clean],
                "quantity": item.get("quantity", 1),
                "unit": item.get("unit", "ชิ้น"),
                "pricePerUnit": item.get("pricePerUnit", 0),
                "totalPrice": item.get("totalPrice", 0),
            })
            continue

        # semantic + dimension-near
        best = None
        for rn, orig in ref_norm_map.items():
            dims_r = ref_dims_map[rn]
            if _dims_close(dims_t, dims_r) and _semantic_compatible(nm, orig):
                best = orig
                break

        if best:
            matched.append({
                "name": best,
                "quantity": item.get("quantity", 1),
                "unit": item.get("unit", "ชิ้น"),
                "pricePerUnit": item.get("pricePerUnit", 0),
                "totalPrice": item.get("totalPrice", 0),
            })
        else:
            uniques.append(item)

    return {"matchedItems": matched, "uniqueItems": uniques}

# =========================================
# =========== Excel I/O Functions =========
# =========================================

def open_or_create_excel(path: str):
    if path and os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
    ws = wb.active
    return wb, ExcelSheetAdapter(ws)

def ensure_first_three_rows_exist(ws_adapter: ExcelSheetAdapter):
    p = []
    for i in range(1, 4):
        p.append({"range": f"A{i}:B{i}", "values": [["", ""]]})
    ws_adapter.batch_update(p, value_input_option="USER_ENTERED")

def _last_non_empty_col_in_top_rows(ws_adapter: ExcelSheetAdapter):
    vals = ws_adapter.get_all_values()
    last = ITEM_MASTER_LIST_COL
    for row in vals[:HEADER_ROW]:
        for i, c in enumerate(row, start=1):
            if str(c).strip():
                last = max(last, i)
    return last

def find_next_available_column(ws_adapter: ExcelSheetAdapter):
    start_col = ITEM_MASTER_LIST_COL + 1
    last_used = _last_non_empty_col_in_top_rows(ws_adapter)
    if last_used < start_col:
        return start_col
    offset = last_used - start_col + 1
    groups_used = (offset + COLUMNS_PER_SUPPLIER - 1) // COLUMNS_PER_SUPPLIER
    return start_col + groups_used * COLUMNS_PER_SUPPLIER

# =========================================
# =========== UPDATE to Excel =============
# =========================================

def update_excel_for_single_file(ws_adapter: ExcelSheetAdapter, data):
    ensure_first_three_rows_exist(ws_adapter)
    start_row = HEADER_ROW + 1
    sheet_values = ws_adapter.get_all_values()
    existing_products = []
    summary_row_map = {}
    first_summary_row = -1

    for row_idx, row in enumerate(sheet_values[HEADER_ROW:], start=start_row):
        if len(row) >= ITEM_MASTER_LIST_COL and row[ITEM_MASTER_LIST_COL - 1].strip():
            cell_value = row[ITEM_MASTER_LIST_COL - 1].strip()
            if cell_value in SUMMARY_LABELS:
                if first_summary_row == -1:
                    first_summary_row = row_idx
                summary_row_map[cell_value] = row_idx
            else:
                product_name = clean_product_name(cell_value)
                existing_products.append({"name": product_name, "row": row_idx})

    existing_suppliers = {}
    header_row_values = sheet_values[COMPANY_NAME_ROW - 1] if sheet_values else []
    for col_idx in range(
        ITEM_MASTER_LIST_COL + 1,
        len(header_row_values) + 1,
        COLUMNS_PER_SUPPLIER,
    ):
        supplier_name = ""
        if COMPANY_NAME_ROW - 1 < len(sheet_values) and (col_idx - 1) < len(sheet_values[COMPANY_NAME_ROW - 1]):
            supplier_name = sheet_values[COMPANY_NAME_ROW - 1][col_idx - 1].strip()
        if supplier_name:
            existing_suppliers[supplier_name] = col_idx

    next_avail_col = find_next_available_column(ws_adapter)

    products = data.get("products", [])
    if not products:
        return 0

    for product in products:
        if product.get("name"):
            product["name"] = clean_product_name(product["name"])

    company_name = data.get("company", "Unknown Company")
    col_idx = existing_suppliers.get(company_name, next_avail_col)
    if col_idx == next_avail_col:
        next_avail_col += COLUMNS_PER_SUPPLIER

    batch_requests = [
        {
            "range": f"{get_column_letter(col_idx)}{COMPANY_NAME_ROW}",
            "values": [[company_name]],
        },
        {
            "range": f"{get_column_letter(col_idx)}{CONTACT_INFO_ROW}",
            "values": [[f"{data.get('contact','')}".strip()]],
        },
        {
            "range": f"{get_column_letter(col_idx)}{HEADER_ROW}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{HEADER_ROW}",
            "values": [["ปริมาณ", "หน่วย", "ราคาต่อหน่วย", "รวมเป็นเงิน"]],
        },
    ]

    reference_data = [{"name": item["name"]} for item in existing_products]
    match_results = match_products_smart(products, reference_data)
    matched_items = match_results["matchedItems"]
    unique_items = match_results["uniqueItems"]

    populated_rows = set()
    for item in matched_items:
        item_name = item.get("name", "")
        for existing in existing_products:
            if existing["name"] == item_name and existing["row"] not in populated_rows:
                batch_requests.append(
                    {
                        "range": f"{get_column_letter(col_idx)}{existing['row']}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{existing['row']}",
                        "values": [[
                            item.get("quantity", 1),
                            item.get("unit", "ชิ้น"),
                            item.get("pricePerUnit", 0),
                            item.get("totalPrice", 0),
                        ]],
                    }
                )
                populated_rows.add(existing["row"])
                break

    new_products = []
    for item in unique_items:
        if isinstance(item, dict) and "name" in item:
            item["name"] = clean_product_name(item["name"])
            if not any(existing["name"] == item["name"] for existing in existing_products):
                new_products.append(item)

    insertion_row = first_summary_row if first_summary_row > 0 else (start_row + len(existing_products))

    if new_products:
        ws_adapter.insert_rows([[""]]*len(new_products), insertion_row)
        row_shift = len(new_products)
        if first_summary_row > 0:
            for label in list(summary_row_map.keys()):
                summary_row_map[label] += row_shift

        for i, product in enumerate(new_products):
            row = insertion_row + i
            batch_requests.append(
                {
                    "range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}",
                    "values": [[product.get("name", "Unknown Product")]],
                }
            )
            batch_requests.append(
                {
                    "range": f"{get_column_letter(col_idx)}{row}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{row}",
                    "values": [[
                        product.get("quantity", 1),
                        product.get("unit", "ชิ้น"),
                        product.get("pricePerUnit", 0),
                        product.get("totalPrice", 0),
                    ]],
                }
            )

    price_col = col_idx + COLUMNS_PER_SUPPLIER - 1
    summary_items = [
        ("รวมเป็นเงิน", data.get("totalPrice", 0)),
        ("ภาษีมูลค่าเพิ่ม 7%", data.get("totalVat", 0)),
        ("ยอดรวมทั้งสิ้น", data.get("totalPriceIncludeVat", 0)),
        ("กำหนดยืนราคา (วัน)", data.get("priceGuaranteeDay", "")),
        ("ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO", data.get("deliveryTime", "")),
        ("การชำระเงิน", data.get("paymentTerms", "")),
        ("อื่น ๆ", data.get("otherNotes", "")),
    ]

    if summary_row_map:
        for label, value in summary_items:
            if label in summary_row_map:
                batch_requests.append(
                    {
                        "range": f"{get_column_letter(price_col)}{summary_row_map[label]}",
                        "values": [[value]],
                    }
                )
    else:
        summary_row = insertion_row + len(new_products) + 2
        for i, (label, value) in enumerate(summary_items):
            row = summary_row + i
            batch_requests.append(
                {
                    "range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}",
                    "values": [[label]],
                }
            )
            batch_requests.append(
                {"range": f"{get_column_letter(price_col)}{row}", "values": [[value]]}
            )

    if batch_requests:
        ws_adapter.batch_update(batch_requests, value_input_option="USER_ENTERED")

    return 1

def update_excel_with_multiple_files(ws_adapter: ExcelSheetAdapter, all_json_data):
    if len(all_json_data) == 1:
        return update_excel_for_single_file(ws_adapter, all_json_data[0])

    ensure_first_three_rows_exist(ws_adapter)
    start_row = HEADER_ROW + 1
    sheet_values = ws_adapter.get_all_values()
    existing_products = []

    for row_idx, row in enumerate(sheet_values[HEADER_ROW:], start=start_row):
        if len(row) >= ITEM_MASTER_LIST_COL and row[ITEM_MASTER_LIST_COL - 1].strip():
            cell_value = row[ITEM_MASTER_LIST_COL - 1].strip()
            if cell_value not in SUMMARY_LABELS:
                product_name = clean_product_name(cell_value)
                existing_products.append({"name": product_name, "row": row_idx})

    existing_suppliers = {}
    header_row_values = sheet_values[COMPANY_NAME_ROW - 1] if sheet_values else []
    for col_idx in range(
        ITEM_MASTER_LIST_COL + 1,
        len(header_row_values) + 1,
        COLUMNS_PER_SUPPLIER,
    ):
        supplier_name = ""
        if (col_idx - 1) < len(header_row_values):
            supplier_name = header_row_values[col_idx - 1].strip()
        if supplier_name:
            existing_suppliers[supplier_name] = col_idx

    next_avail_col = find_next_available_column(ws_adapter)

    for data in all_json_data:
        products = data.get("products", [])
        if not products:
            continue

        for product in products:
            if product.get("name"):
                product["name"] = clean_product_name(product["name"])

        company_name = data.get("company", "Unknown Company")
        col_idx = existing_suppliers.get(company_name, next_avail_col)
        if col_idx == next_avail_col:
            next_avail_col += COLUMNS_PER_SUPPLIER

        batch_requests = [
            {
                "range": f"{get_column_letter(col_idx)}{COMPANY_NAME_ROW}",
                "values": [[company_name]],
            },
            {
                "range": f"{get_column_letter(col_idx)}{CONTACT_INFO_ROW}",
                "values": [[f"{data.get('contact','')}".strip()]],
            },
            {
                "range": f"{get_column_letter(col_idx)}{HEADER_ROW}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{HEADER_ROW}",
                "values": [["ปริมาณ", "หน่วย", "ราคาต่อหน่วย", "รวมเป็นเงิน"]],
            },
        ]

        reference_data = [{"name": item["name"]} for item in existing_products]
        match_results = match_products_smart(products, reference_data)
        matched_items = match_results["matchedItems"]
        unique_items = match_results["uniqueItems"]

        populated_rows = set()

        for item in matched_items:
            item_name = item.get("name", "")
            for existing in existing_products:
                if existing["name"] == item_name and existing["row"] not in populated_rows:
                    batch_requests.append(
                        {
                            "range": f"{get_column_letter(col_idx)}{existing['row']}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{existing['row']}",
                            "values": [[
                                item.get("quantity", 1),
                                item.get("unit", "ชิ้น"),
                                item.get("pricePerUnit", 0),
                                item.get("totalPrice", 0),
                            ]],
                        }
                    )
                    populated_rows.add(existing["row"])
                    break

        new_products = []
        for item in unique_items:
            if isinstance(item, dict) and "name" in item:
                item["name"] = clean_product_name(item["name"])
                if not any(existing["name"] == item["name"] for existing in existing_products):
                    new_products.append(item)

        next_row = start_row + len(existing_products)

        if new_products:
            ws_adapter.insert_rows([[""]]*len(new_products), next_row)

            for i, product in enumerate(new_products):
                row = next_row + i
                batch_requests.append(
                    {
                        "range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}",
                        "values": [[product.get("name", "Unknown Product")]],
                    }
                )
                batch_requests.append(
                    {
                        "range": f"{get_column_letter(col_idx)}{row}:{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{row}",
                        "values": [[
                            product.get("quantity", 1),
                            product.get("unit", "ชิ้น"),
                            product.get("pricePerUnit", 0),
                            product.get("totalPrice", 0),
                        ]],
                    }
                )
                existing_products.append({"name": product.get("name", "Unknown Product"), "row": row})

        summary_row = start_row + len(existing_products) + 2
        summary_items = [
            ("รวมเป็นเงิน", data.get("totalPrice", 0)),
            ("ภาษีมูลค่าเพิ่ม 7%", data.get("totalVat", 0)),
            ("ยอดรวมทั้งสิ้น", data.get("totalPriceIncludeVat", 0)),
            ("กำหนดยืนราคา (วัน)", data.get("priceGuaranteeDay", "")),
            ("ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO", data.get("deliveryTime", "")),
            ("การชำระเงิน", data.get("paymentTerms", "")),
            ("อื่น ๆ", data.get("otherNotes", "")),
        ]

        for i, (label, value) in enumerate(summary_items):
            row = summary_row + i
            batch_requests.append(
                {"range": f"{get_column_letter(ITEM_MASTER_LIST_COL)}{row}", "values": [[label]]}
            )
            batch_requests.append(
                {"range": f"{get_column_letter(col_idx+COLUMNS_PER_SUPPLIER-1)}{row}", "values": [[value]]}
            )

        if batch_requests:
            ws_adapter.batch_update(batch_requests, value_input_option="USER_ENTERED")

        if company_name not in existing_suppliers:
            existing_suppliers[company_name] = col_idx

    return len(all_json_data)

# =========================================
# =========== File Type & Process =========
# =========================================

def get_file_type(file_path):
    mime_type, _ = mimetypes.guess_type(file_path)
    if mime_type:
        if mime_type.startswith("image/"):
            return "image"
        elif mime_type == "application/pdf":
            return "pdf"
        elif mime_type in [
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ]:
            return "word"
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".gif", ".webp"]:
        return "image"
    elif ext == ".pdf":
        return "pdf"
    elif ext in [".doc", ".docx"]:
        return "word"
    return "unknown"

def _wait_for_file_active(uploaded_file, timeout=180, poll=1.0):
    start = time.time()
    name = getattr(uploaded_file, "name", None)
    if not name:
        return uploaded_file
    while time.time() - start < timeout:
        try:
            f2 = genai.get_file(name)
            state = getattr(f2, "state", None)
            if state == "ACTIVE":
                return f2
            time.sleep(poll)
        except Exception:
            time.sleep(poll)
    return uploaded_file

def process_file(file_path):
    file_name = os.path.basename(file_path)
    with open(file_path, "rb") as src:
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file_name)[1]) as tmp_file:
            tmp_file.write(src.read())
            tmp_file_path = tmp_file.name

    uploaded_gemini_file = genai.upload_file(path=tmp_file_path, display_name=file_name)
    uploaded_gemini_file = _wait_for_file_active(uploaded_gemini_file)

    file_type = get_file_type(file_path)
    prompt_to_use = image_prompt if file_type == "image" else prompt

    model_flash = genai.GenerativeModel(model_name="gemini-2.5-flash", generation_config={"temperature": 0.1, "top_p": 0.95})
    resp = model_flash.generate_content([prompt_to_use, uploaded_gemini_file])
    d = extract_json_from_text(getattr(resp, "text", "") or "")

    if not d or not d.get("products"):
        model_pro = genai.GenerativeModel(model_name="gemini-2.5-pro", generation_config={"temperature": 0.1, "top_p": 0.95})
        resp_pro = model_pro.generate_content([prompt_to_use, uploaded_gemini_file])
        d = extract_json_from_text(getattr(resp_pro, "text", "") or "")

    d = validate_json_data(d) if d else None
    d = enhance_with_gemini(d) if d else None

    if d:
        result = {"file_name": file_name, "data": d}
    else:
        result = {"file_name": file_name, "error": "Failed to extract structured data from the document after multiple attempts."}

    if tmp_file_path and os.path.exists(tmp_file_path):
        os.unlink(tmp_file_path)
    if uploaded_gemini_file and getattr(uploaded_gemini_file, "name", None):
        genai.delete_file(uploaded_gemini_file.name)
    return result

def process_files(file_paths, excel_path=DEFAULT_EXCEL_PATH):
    data_list = []
    error_list = []
    total_files = len(file_paths)
    with st.status(f"Processing {total_files} files...", expanded=True) as status:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(total_files, 10)) as executor:
            future_to_file = {executor.submit(process_file, path): path for path in file_paths}
            processed_count = 0
            for future in concurrent.futures.as_completed(future_to_file):
                processed_count += 1
                result = future.result()
                if "error" in result and result["error"]:
                    st.warning(f"⚠️ Failed to process {result['file_name']}: {result['error']}")
                    error_list.append(result)
                else:
                    st.write(f"✓ Successfully processed {result['file_name']}")
                    data_list.append(result["data"])
                status.progress(processed_count / total_files, text=f"Processed {processed_count}/{total_files} files")

    if data_list:
        status.update(label="Updating Excel file...", state="running")
        wb, ws_adapter = open_or_create_excel(excel_path)
        update_excel_with_multiple_files(ws_adapter, data_list)
        wb.save(excel_path)
        status.update(label=f"Saved to {excel_path}", state="complete", expanded=False)
    elif not error_list:
        status.update(label="No data could be extracted from the files.", state="complete")
    else:
        status.update(label="Processing finished with errors.", state="error", expanded=True)
    return data_list, error_list

def process_pdfs(pdf_paths, excel_path=DEFAULT_EXCEL_PATH):
    return process_files(pdf_paths, excel_path)

# =========================================
# =========== PROMPTS (เต็มตามเดิม) =======
# =========================================

prompt = """# System Message for Product List Extraction (PDF/Text Table Processing)
## CRITICAL: ANTI-HALLUCINATION WARNING
You MUST ONLY extract information that is EXPLICITLY visible in the document. 
- DO NOT add data from anywhere other than the one in the uploaded document. Adding data that is not from the document is a serious mistake.
- DO NOT create or invent any products, prices, or specifications
- DO NOT add data from the attached Example in the prompt.
- DO NOT add products that are not clearly listed as distinct line items
- DO NOT attempt to break down a single product into multiple products
- DO NOT interpret descriptive text as separate products
- If uncertain about any information, LEAVE IT OUT rather than guessing

## CRITICAL: CONTACT INFORMATION EXTRACTION
Pay special attention to contact information in the document header or footer:
- Extract any email addresses (example@domain.com)
- Extract any phone numbers (formats like 02-3384825, 081-1234567, 095-525-2623)
- Put email and phone details in the "contact" field
- Format: "Email: email@example.com, Phone: 081-234-5678"
- Use only the phone number and email address on the letterhead. Do not add "บริษัท ลูก้า แอสเซท จำกัด, 081-781-7283" contact information

## CRITICAL: COMPLETE EXTRACTION REQUIREMENT
You MUST extract ALL products visible in the document:
- Extract EVERY product line item visible in the document
- Preserve hierarchical structure (groups/categories) of products if present
- Ensure NO products are missed or skipped
- Each row with a distinct price is one product

## Input Format
Provide PDF files (or images/tables with text extraction) containing product information (receipts, invoices, product lists, etc.).

## Task
Extract ALL product information EXPLICITLY visible in the document:
- Extract products with quantities, units, and prices
- Preserve parent-child relationships (main categories and sub-items)
- Maintain hierarchical product groupings (numbered sections, categories)
- Also extract additional quotation details like price validity, delivery time, payment terms, etc.

## Hierarchical Structure Handling
Many quotations organize products hierarchically. When you see this:
- Include category names in product descriptions (e.g., "งานบันไดกระจก งานพื้นตก - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม.")
- If product descriptions begin with numbers (1, 2, 3...), REMOVE those numbers
- Include all parent category information in each product's name without the leading numbers

## Output Format (JSON only)
You must return ONLY this JSON structure:
{
  "company": "company name or first name + last name (NEVER null)",
  "vat": true,
  "name": "customer name or null",
  "contact": "phone number or email or null",
  "priceGuaranteeDay": 30
  "deliveryTime": "",
  "paymentTerms": "",
  "otherNotes": "",
  "products": [
    {
      "name": "full product description including ALL parent category info, specifications AND dimensions WITHOUT leading numbers",
      "quantity": 1,
      "unit": "match the unit shown in the document (e.g., แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, ชุด)",
      "pricePerUnit": 0,
      "totalPrice": 0
    }
  ],
  "totalPrice": 0,
  "totalVat": 0,
  "totalPriceIncludeVat": 0
}

## Example 1 (Format with Item/ART.No./Description/Qty/Unit/Price columns):
| Item | ART.No. | Description | Qty | Unit | Standard Price | Discount Price | Amount |
|------|---------|-------------|-----|------|----------------|---------------|--------|
| 1    | CPW-xxxx| SPC ลายไม้ 4.5 มิล (ก้างปลา) | 1.00 | ตร.ม. |  | 520.00 | 520.000 |
| 2    |         | ค่าแรงติดตั้ง | 1.00 | ตร.ม. |  | 150.00 | 150.000 |

## Example 2 (Format with ลำดับ/รหัสสินค้า/รายละเอียดสินค้า columns):
| ลำดับ | รหัสสินค้า | รายละเอียดสินค้า | หน่วย | จำนวน | ราคา/หน่วย(บาท) | จำนวนเงิน(บาท) |
|------|---------|----------------|------|------|--------------|------------|
| 1    |         | พื้นไม้ไวนิลลายไม้ปลา 4.5 มม. LKT 4.5 mm x 0.3 mm สีฟ้าเซอร์คูลี (1 กล่อง บรรจุ 18 แผ่น หรือ 1.3 ตร.ม) | ตร.ม. | 1.30 | 680.00 | 884.00 |

## Field Extraction Guidelines

### name (Product Description)
* CRITICAL: Include ALL hierarchical information in each product name:
  - Category names/headings (e.g., "งานบันไดกระจก งานพื้นตก")
  - Sub-category information (e.g., "เหล็กตัวซีชุบสังกะสี")
  - Glass type, thickness (e.g., "กระจกเทมเปอร์ใส หนา 10 มม.")
  - Exact dimensions (e.g., "ขนาด 4.672×0.97 ม.")
* REMOVE any leading numbers (1., 2., 3.) from the product descriptions
* Format hierarchical products as: "[Category Name] - [Material] - [Type] - [Dimensions]"
* Include: ALL distinguishing characteristics that make each product unique
* Example: "งานบันไดกระจก งานพื้นตก - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม."

### unit and quantity (DIRECT EXTRACTION RULE)
* Extract unit and quantity DIRECTLY from each line item as shown
* Use the exact unit shown in the document (ชุด, แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, etc.)
* Extract the exact quantity shown for each product (never assume or calculate)
* NEVER create quantities or units that aren't explicitly shown in the document
* Pay special attention to decimal quantities - extract the full decimal precision

### pricePerUnit and totalPrice
* Extract ONLY prices clearly visible in the document
* Use numeric values only (no currency symbols)
* Extract cleanly from pricing fields as shown in each line item
* NEVER calculate or estimate prices that aren't explicitly shown
* NEVER combine different products' prices
* Pay special attention to decimal prices - extract the EXACT decimal values shown

### Additional Quotation Details
* Extract these additional fields if present:
  - "กำหนดยืนราคา (วัน)", "กำหนดยืนราคา", "การยืนราคา" - Price validity period in days (priceGuaranteeDay)
  - "ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO" - Delivery time after PO (deliveryTime)
  - "การชำระเงิน" - Payment terms (paymentTerms)
  - "อื่น ๆ" - Other notes (otherNotes)
* Extract as text exactly as written, preserving numbers and Thai language

### CRITICAL: Pricing summaries and summary values
* Extract the exact values for these three summary items:
  - "รวมเป็นเงิน" - the initial subtotal (totalPrice)
  - "ภาษีมูลค่าเพิ่ม 7%" - the VAT amount (totalVat)
  - "ยอดรวมทั้งสิ้น" - the final total (totalPriceIncludeVat)
* Alternative labels to match:
  - For totalPrice: "รวม", "รวมเป็นเงิน", "ราคารวม", "Total", "TOTAL AMOUNT", "รวมราคา"
  - For totalVat: "ภาษีมูลค่าเพิ่ม 7%", "VAT 7%"
  - For totalPriceIncludeVat: "ยอดรวมทั้งสิ้น", "รวมทั้งหมด", "รวมเงินทั้งสิน", "ราคารวมสุทธิ", "รวมราคางานทั้งหมดตามสัญญา"
* Extract the exact values as shown (remove commas, currency symbols)
* CRITICAL: Preserve full decimal precision in all monetary values

## FINAL VERIFICATION
Review the extracted products one last time and verify:
1. Count the number of products you've extracted
2. Verify this matches EXACTLY with the number of product rows visible in the document
3. Check that ALL products have proper hierarchical information included WITHOUT leading numbers
4. Ensure NO products are missing - every line item with a price must be extracted
5. Confirm all dimensions and specifications are preserved correctly
6. Verify all decimal values (quantities and prices) maintain their full precision
"""

image_prompt = """# System Message for Product List Extraction from Images
## CRITICAL: ANTI-HALLUCINATION WARNING
You MUST ONLY extract information that is EXPLICITLY visible in the image. 
- DO NOT create or invent any products, prices, or specifications
- DO NOT add products that are not clearly listed as distinct line items
- DO NOT interpret descriptive text as separate products
- If text is unclear or unreadable, mark it as uncertain rather than guessing

## CRITICAL: COMPLETE EXTRACTION REQUIREMENT
You MUST extract ALL products visible in the image:
- Extract EVERY product line item visible in the image
- Preserve hierarchical structure (groups/categories) of products if present
- Ensure NO products are missed or skipped
- Each row with a distinct price is one product

## Input Format
I'm providing an image of a document containing product information.

## Task
Extract ALL product information EXPLICITLY visible in the image:
- Extract products with quantities, units, and prices
- Preserve parent-child relationships (main categories and sub-items)
- Maintain hierarchical product groupings (numbered sections, categories)
- Also extract additional quotation details like price validity, delivery time, payment terms, etc.

## Hierarchical Structure Handling
Many quotations organize products hierarchically. When you see this:
- Include category names in product descriptions (e.g., "งานบันไดกระจก งานพื้นตก - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม.")
- If product descriptions begin with numbers (1, 2, 3...), REMOVE those numbers
- Include all parent category information in each product's name without the leading numbers

## Output Format (JSON only)
You must return ONLY this JSON structure:
{
  "company": "company name or first name + last name (NEVER null)",
  "vat": true,
  "contact": "phone number or email or null",
  "priceGuaranteeDay": 30
  "deliveryTime": "",
  "paymentTerms": "",
  "otherNotes": "",
  "products": [
    {
      "name": "full product description including ALL parent category info, specifications AND dimensions WITHOUT leading numbers",
      "quantity": 1,
      "unit": "match the unit shown in the document (e.g., แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, ชุด)",
      "pricePerUnit": 0,
      "totalPrice": 0
    }
  ],
  "totalPrice": 0,
  "totalVat": 0,
  "totalPriceIncludeVat": 0
}


## Example 1 (Format with Item/ART.No./Description/Qty/Unit/Price columns):
| Item | ART.No. | Description | Qty | Unit | Standard Price | Discount Price | Amount |
|------|---------|-------------|-----|------|----------------|---------------|--------|
| 1    | CPW-xxxx| SPC ลายไม้ 4.5 มิล (ก้างปลา) | 1.00 | ตร.ม. |  | 520.00 | 520.000 |
| 2    |         | ค่าแรงติดตั้ง | 1.00 | ตร.ม. |  | 150.00 | 150.000 |


## Example 2 (Format with ลำดับ/รหัสสินค้า/รายละเอียดสินค้า columns):
| ลำดับ | รหัสสินค้า | รายละเอียดสินค้า | หน่วย | จำนวน | ราคา/หน่วย(บาท) | จำนวนเงิน(บาท) |
|------|---------|----------------|------|------|--------------|------------|
| 1    |         | พื้นไม้ไวนิลลายไม้ปลา 4.5 มม. LKT 4.5 mm x 0.3 mm สีฟ้าเซอร์คูลี (1 กล่อง บรรจุ 18 แผ่น หรือ 1.3 ตร.ม) | ตร.ม. | 1.30 | 680.00 | 884.00 |


## Field Extraction Guidelines

### name (Product Description)
* CRITICAL: Include ALL hierarchical information in each product name:
  - Category names/headings (e.g., "งานบันไดกระจก งานพื้นตก")
  - Sub-category information (e.g., "เหล็กตัวซีชุบสังกะสี")
  - Glass type, thickness (e.g., "กระจกเทมเปอร์ใส หนา 10 มม.")
  - Exact dimensions (e.g., "ขนาด 4.672×0.97 ม.")
* REMOVE any leading numbers (1., 2., 3.) from the product descriptions
* Format hierarchical products as: "[Category Name] - [Material] - [Type] - [Dimensions]"
* Include: ALL distinguishing characteristics that make each product unique
* Example: "งานบันไดกระจก งานพื้นตก - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม."

### unit and quantity (DIRECT EXTRACTION RULE)
* Extract unit and quantity DIRECTLY from each line item as shown
* Use the exact unit shown in the document (ชุด, แผ่น, ตร.ม., ชิ้น, ตัว, เมตร, จำนวนต่อชุด etc.)
* Extract the exact quantity shown for each product (never assume or calculate)
* NEVER create quantities or units that aren't explicitly shown in the document
* Pay special attention to decimal quantities - extract the full decimal precision

### pricePerUnit and totalPrice
* Extract ONLY prices clearly visible in the document
* Use numeric values only (no currency symbols)
* Extract cleanly from pricing fields as shown in each line item
* NEVER calculate or estimate prices that aren't explicitly shown
* NEVER combine different products' prices
* Pay special attention to decimal prices - extract the EXACT decimal values shown

### Additional Quotation Details
* Extract these additional fields if present:
  - "กำหนดยืนราคา (วัน)", "กำหนดยืนราคา", "การยืนราคา" - Price validity period in days (priceGuaranteeDay)
  - "ระยะเวลาส่งมอบสินค้าหลังจากได้รับ PO" - Delivery time after PO (deliveryTime)
  - "การชำระเงิน" - Payment terms (paymentTerms)
  - "อื่น ๆ" - Other notes (otherNotes)
* Extract as text exactly as written, preserving numbers and Thai language

### CRITICAL: Pricing summaries and summary values
* Extract the exact values for these three summary items:
  - "รวมเป็นเงิน" - the initial subtotal (totalPrice)
  - "ภาษีมูลค่าเพิ่ม 7%" - the VAT amount (totalVat)
  - "ยอดรวมทั้งสิ้น" - the final total (totalPriceIncludeVat)
* Alternative labels to match:
  - For totalPrice: "รวม", "รวมเป็นเงิน", "ราคารวม", "Total", "TOTAL AMOUNT", "รวมราคา"
  - For totalVat: "ภาษีมูลค่าเพิ่ม 7%", "VAT 7%"
  - For totalPriceIncludeVat: "ยอดรวมทั้งสิ้น", "รวมทั้งหมด", "รวมเงินทั้งสิน", "ราคารวมสุทธิ", "รวมราคางานทั้งหมดตามสัญญา"
* Extract the exact values as shown (remove commas, currency symbols)
* CRITICAL: Preserve full decimal precision in all monetary values

## FINAL VERIFICATION
Review the extracted products one last time and verify:
1. Count the number of products you've extracted
2. Verify this matches EXACTLY with the number of product rows visible in the image
3. Check that ALL products have proper hierarchical information included WITHOUT leading numbers
4. Ensure NO products are missing - every line item with a price must be extracted
5. Confirm all dimensions and specifications are preserved correctly
6. Verify all decimal values (quantities and prices) maintain their full precision
"""

validation_prompt = """
You are a data validation expert specializing in Thai construction quotations.
I've extracted product data from a document, but there may be missing products or hierarchical relationships.

## CRITICAL: COMPLETE DATA CHECK
Your primary task is to ensure ALL products are correctly extracted with their hierarchical structure:
1. Check that all products visible in the document have been extracted
2. Ensure parent-child relationships and category groupings are preserved
3. Verify that all products have complete descriptions including their category name
4. Make sure no products are missing dimensions or specifications

## CRITICAL: PRESERVE PRODUCT HIERARCHY
Thai construction quotations often organize products hierarchically by categories:
- Category names with descriptive details
- Materials and specifications
- Dimensions

Each product must include its complete hierarchy:
"[Category Name] - [Material] - [Type] - [Dimensions]"

Examples (DO NOT add data from the attached Example in the prompt): 
- "งานบันไดกระจก งานพื้นตก - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - กระจกเทมเปอร์ใส หนา 10 มม. ขนาด 4.672×0.97 ม."
- "งานพื้นตก (ชั้นลอย) - เหล็กตัวซีชุบสังกะสี ไม่รวมปูน - เทมเปอร์ใส หนา 10 มม. ขนาด 3.565×0.97 ม."

## CRITICAL: DECIMAL NUMBER ACCURACY
Pay special attention to:
1. Quantities with decimals (extract full precision)
2. Dimensions with decimals (preserve exact measurements)
3. Prices with decimals (maintain exact values)

## CRITICAL: CLEAN PRODUCT DESCRIPTIONS
1. REMOVE any leading numbers (1., 2., 3., etc.) from product descriptions
2. Ensure NO product descriptions begin with numbering
3. Maintain all other hierarchical information and details

Review the data carefully and FIX these issues:
1. ADD any missing products that should be extracted from the source document
2. FIX product names to include complete hierarchical information WITHOUT leading numbers
3. ENSURE all dimensions and specifications are preserved with full decimal precision
4. VERIFY every product has the correct quantity, unit, price and total with full decimal precision

Original extraction:
{extracted_json}

Return ONLY a valid JSON object with no explanations.
"""

matching_prompt = """
You are a meticulous data architect specializing in product ontology for construction and home appliance materials. Your primary mission is to analyze product lists from different suppliers, establish a single "canonical" master product name for each item, and then map all supplier variations to that canonical name.
Your logic must be hierarchical and rule-based. Follow this algorithm precisely.

### **Core Objective: Create and Match to a Canonical Name**

The "Canonical Name" is the single source of truth for a product. You must construct it using this strict format:
**`[Group] - [Normalized Product Type] - [Primary Model/Identifier]`**

-   **`[Group]`**: The project phase or room size (e.g., "1BR+2BR(57-70sqm.)", "2BR (90sqm.)"). This is the **highest-priority** matching key.
-   **`[Normalized Product Type]`**: The generic category of the product (e.g., "Hood", "Induction Hob", "Sink"). You must deduce this from various descriptions.
-   **`[Primary Model/Identifier]`**: The most specific model number available (e.g., "EL 60", "MWE 255 FI").

### **Mandatory 4-Step Matching Algorithm**

For every product you process, you must follow these steps in order:

#### **Step 1: Group Matching (Non-Negotiable Filter)**
-   This is the most critical step. Products can **ONLY** be considered a match if they belong to the **exact same `[Group]`**.
-   Example: A "Hood" from "1BR+2BR(57-70sqm.)" can **NEVER** match a "Hood" from "2BR (90sqm.)". They are distinct line items.
-   Recognize semantic equivalents for groups, e.g., "1 BEDROOM" is the same as "1BR+2BR(57-70sqm.)".

#### **Step 2: Product Type Normalization & Keyword Mapping**
-   After filtering by group, identify the core product type. You must normalize different supplier descriptions into one standard type.
-   Use this keyword map as your guide:
    -   **"Hood"**: `Slimline Hood`, `BI telescopic hood`, `HOOD PIAVE 60 XS`
    -   **"Induction Hob"**: `Induction Hob`, `Hob Electric`, `HOB INDUCTION`
    -   **"Microwave"**: `Built-in Microwave`, `Microwave Oven`, `MICROWAVE FMWO 25 NH I`
    -   **"Sink"**: `Undermount Sink`, `Sink Stainless Steel`, `SINK BXX 210-45`
    -   **"Tap"**: `Sink Single Tap`, `Tap`, `TAP LANNAR`

#### **Step 3: Specification & Model Analysis**
-   Once Group and Normalized Type match, use the model number and other specifications (`Model`, `Description` columns) to create the full canonical name and to confirm the match.
-   The model number itself does not have to be identical between suppliers if the Group and Normalized Type are a clear match. The model number's purpose is to create the *unique canonical name* for that row.

#### **Step 4: Construct Final Output**
-   Based on the matches found, generate the final JSON.

### **Critical Rules & Constraints**

1.  **Group is King:** If the group doesn't match, nothing else matters.
2.  **Type over Model:** A strong match on `Group` + `Normalized Product Type` is more important than a weak match on `Model` number.
3.  **One-to-One Mapping:** A reference item (a canonical name you create) can only be matched once per supplier list.
4.  **No Imagination:** Only use information explicitly present in the data. If you cannot confidently normalize a product type, classify it as unique.

### **Walkthrough Example: Matching "Hoods"**

**Goal:** Match the first item from all three suppliers.

1.  **Teka:**
    -   **Input:** Group=`1BR+2BR(57-70sqm.)`, Model=`EL 60`, Desc=`Slimline Hood`
    -   **Analysis:** Group is "1BR...". Type normalizes from "Slimline Hood" to **"Hood"**. Model is `EL 60`.
    -   **Canonical Name Created:** `1BR+2BR(57-70sqm.) - Hood - EL 60`

2.  **Hisense (Gorenje):**
    -   **Input:** Group=`1BR+2BR (57-70Sqm.)`, Product=`TH62E3X`, Desc=`BI telescopic hood...`
    -   **Analysis:** Group is "1BR...". It matches Teka's group. Type normalizes from "BI telescopic hood" to **"Hood"**. It matches the normalized type.
    -   **Conclusion:** This is a match for the same row.

3.  **Franke:**
    -   **Input:** Group=`1 BEDROOM`, Product Category=`Hood`, Mode=`PIAVE 60 XS`
    -   **Analysis:** Group "1 BEDROOM" is semantically identical to "1BR...". It matches. Type is explicitly `Hood`. It matches.
    -   **Conclusion:** This is also a match for the same row.

All three products are mapped to the canonical name `1BR+2BR(57-70sqm.) - Hood - EL 60`, and their respective data will be aligned on this single row in the final output.

### **Input & Output Format**

-   **Input:** `target_products` (from a new quotation) and `reference_products` (the existing master list of canonical names).
-   **Output:** You **MUST** return a JSON object with this exact structure:

{{
  "matchedItems": [
    {{
      "name": "The canonical reference name this product matched to.",
      "quantity": "target quantity",
      "unit": "target unit",
      "pricePerUnit": "target price per unit",
      "totalPrice": "target total price"
    }}
  ],
  "uniqueItems": [
    {{
      "name": "The full, descriptive name of the target product that could not be matched.",
      "quantity": "target quantity",
      "unit": "target unit",
      "pricePerUnit": "target price per unit",
      "totalPrice": "target total price"
    }}
  ]
}}

## Target Products:
{target_products}

## Reference Products:
{reference_products}
"""

def main():
    st.set_page_config(page_title="ระบบประมวลผลใบเสนอราคา → Excel", layout="centered")
    st.sidebar.title("ตั้งค่า")
    google_api_key = st.sidebar.text_input(
        "Enter your GOOGLE_API_KEY", 
        value=st.session_state.get("google_api_key", ""), 
        type="password",
        key="google_api_key_input"
    )

    if st.sidebar.button("ยืนยัน API Key", key="confirm_api_key", use_container_width=True):
        if google_api_key:
            genai.configure(api_key=google_api_key)
            st.session_state.google_api_key = google_api_key
            st.session_state.api_key_confirmed = True
            st.sidebar.success("API Key ถูกบันทึกและใช้งานได้แล้ว")
        else:
            st.session_state.api_key_confirmed = False
            st.sidebar.warning("กรุณาใส่ API Key ก่อนยืนยัน")

    st.markdown("<h1 style='text-align: center;'>ระบบประมวลผลใบเสนอราคา → Excel</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>อัปโหลดไฟล์ (PDF/รูปภาพ) แล้วบันทึกผลลงไฟล์ Excel ตามฟอร์แมตเดิม พร้อมแมตช์สินค้าด้วยกติกาใหม่ (semantic + ขนาด)</p>", unsafe_allow_html=True)
    st.markdown("---")
    st.subheader("ขั้นตอนที่ 1: อัปโหลดข้อมูลของคุณ")

    excel_path = st.text_input(
        "Excel Path (จะสร้างใหม่ถ้าไม่มี):",
        value=DEFAULT_EXCEL_PATH,
        placeholder="เช่น /mnt/data/quotation_output.xlsx หรือ path ของ template เดิม"
    )

    uploaded_files = st.file_uploader(
        "เลือกไฟล์ PDF หรือรูปภาพ (หลายไฟล์ได้)",
        type=['pdf', 'jpg', 'jpeg', 'png'],
        accept_multiple_files=True
    )
    if uploaded_files:
        if st.button("🚀 เริ่มประมวลผล และ Save to Excel", use_container_width=True, type="primary"):
            if not st.session_state.get("api_key_confirmed"):
                st.error("❌ กรุณาใส่และยืนยัน Google API Key ในแถบด้านข้างก่อนเริ่มประมวลผล")
                return
            file_paths = []
            with st.spinner("กำลังเตรียมไฟล์เพื่ออัปโหลด..."):
                for uploaded_file in uploaded_files:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[1]) as tmp_file:
                        tmp_file.write(uploaded_file.getvalue())
                        file_paths.append(tmp_file.name)

            results, errors = process_files(file_paths, excel_path)
            st.subheader("ขั้นตอนที่ 2: ผลการประมวลผล")
            if results:
                st.success(f"✅ ประมวลผลสำเร็จ {len(results)} ไฟล์ และบันทึกข้อมูลลง Excel: {excel_path}")
                try:
                    with open(excel_path, "rb") as f:
                        st.download_button(
                            "⬇️ ดาวน์โหลด Excel",
                            data=f,
                            file_name=os.path.basename(excel_path),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception:
                    st.info("💾 ไฟล์ถูกบันทึกในเครื่องเซิร์ฟเวอร์/เครื่องรันแอปเรียบร้อยแล้ว")
                for result_data in results:
                    company_name = result_data.get('company', 'Unknown Company')
                    with st.expander(f"📄 ผลลัพธ์จาก: {company_name}"):
                        st.json(result_data)
            if errors:
                st.error(f"❌ พบข้อผิดพลาด {len(errors)} ไฟล์ ไม่สามารถประมวลผลได้")
                for error_info in errors:
                    with st.expander(f"🚨 ข้อผิดพลาดในไฟล์: {error_info['file_name']}"):
                        st.write("**สาเหตุ:**")
                        st.code(error_info['error'], language=None)
            if not results and not errors:
                st.warning("ไม่สามารถประมวลผลไฟล์ใดๆ ได้ กรุณาตรวจสอบไฟล์และลองใหม่อีกครั้ง")

if __name__ == "__main__":
    main()